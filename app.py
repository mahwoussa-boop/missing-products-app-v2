"""
╔══════════════════════════════════════════════════════════════════╗
║   مهووس — مركز التحكم الشامل  v4.5  (Production-Ready)         ║
║   Mahwous Ultimate Control Center                               ║
║   Streamlit · Anthropic Claude · Google CSE · Railway           ║
╚══════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import numpy as np
import io, re, os, json, time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
import anthropic
try:
    from rapidfuzz import process as rf_process, fuzz as rf_fuzz
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE CONFIG                                                    ║
# ╚══════════════════════════════════════════════════════════════════╝
st.set_page_config(
    page_title="مهووس | مركز التحكم",
    page_icon="🌸",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  GLOBAL CSS — Arabic RTL + Gold Theme                          ║
# ╚══════════════════════════════════════════════════════════════════╝
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;600;700;900&display=swap');

/* ── Global Reset & RTL ─────────────────────────── */
*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"], .stApp, .main, section, div {
  font-family: 'Cairo', sans-serif !important;
}
.stApp { background-color: #f5f0e8; }

/* ── Sidebar ─────────────────────────────────────── */
section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #0f0e0d 0%, #1c1610 100%) !important;
  border-left: 1px solid rgba(184,147,58,0.25);
}
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div { color: #e0d0b0 !important; }
section[data-testid="stSidebar"] .stRadio label { color: #c8b080 !important; }

/* ── Top Header ──────────────────────────────────── */
.mhw-header {
  background: linear-gradient(135deg, #0f0e0d 0%, #1c1610 50%, #2a1e08 100%);
  border: 1px solid rgba(184,147,58,0.3);
  border-radius: 14px;
  padding: 16px 24px;
  display: flex;
  align-items: center;
  gap: 16px;
  margin-bottom: 20px;
  box-shadow: 0 8px 32px rgba(0,0,0,0.2);
}
.mhw-header .emblem {
  width: 52px; height: 52px;
  background: linear-gradient(135deg, #b8933a, #e0b84a);
  border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-size: 26px; font-weight: 900; color: #0f0e0d;
  box-shadow: 0 0 24px rgba(184,147,58,0.6);
  flex-shrink: 0;
}
.mhw-header h1 {
  color: #b8933a; font-size: 1.55rem;
  margin: 0; line-height: 1.2;
}
.mhw-header p { color: rgba(255,255,255,0.38); font-size: 0.78rem; margin: 0; }

/* ── Section Title ───────────────────────────────── */
.sec-title {
  display: flex; align-items: center; gap: 10px;
  margin: 22px 0 14px; direction: rtl;
}
.sec-title .bar {
  width: 5px; height: 24px; border-radius: 3px;
  background: linear-gradient(180deg, #b8933a, #e0b84a);
}
.sec-title h3 { margin: 0; font-size: 1.05rem; font-weight: 800; color: #1a1208; }

/* ── Stats Bar ───────────────────────────────────── */
.stats-bar { display: flex; gap: 12px; flex-wrap: wrap; margin: 14px 0; }
.stat-box {
  flex: 1; min-width: 110px;
  background: white;
  border: 1px solid rgba(184,147,58,0.22);
  border-radius: 12px;
  padding: 14px 16px; text-align: center;
  box-shadow: 0 2px 10px rgba(0,0,0,0.05);
}
.stat-box .n  { font-size: 1.9rem; font-weight: 900; color: #b8933a; line-height: 1; }
.stat-box .lb { font-size: 0.73rem; color: #7a6e60; margin-top: 3px; }

/* ── Upload Zone ─────────────────────────────────── */
.upload-zone {
  border: 2px dashed rgba(184,147,58,0.38);
  border-radius: 16px; padding: 2.5rem;
  text-align: center;
  background: rgba(184,147,58,0.035);
  transition: all 0.2s;
}
.upload-zone:hover {
  border-color: #b8933a;
  background: rgba(184,147,58,0.07);
}
.uz-icon  { font-size: 3rem; }
.uz-title { font-size: 1.08rem; font-weight: 800; color: #1a1208; margin: 6px 0 3px; }
.uz-sub   { font-size: 0.8rem; color: #9a8e80; }

/* ── Tool Tabs ───────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
  background: rgba(184,147,58,0.06);
  border-radius: 10px; padding: 4px; gap: 4px;
  border-bottom: none !important;
}
.stTabs [data-baseweb="tab"] {
  border-radius: 8px !important;
  font-family: 'Cairo', sans-serif !important;
  font-weight: 700 !important; font-size: 0.82rem !important;
  padding: 8px 14px !important;
}
.stTabs [aria-selected="true"] {
  background: linear-gradient(135deg, #b8933a, #e0b84a) !important;
  color: #0f0e0d !important;
}

/* ── Alerts ──────────────────────────────────────── */
.al-info {
  background: #e8f4fd; border-right: 4px solid #1976d2;
  border-radius: 8px; padding: 10px 14px;
  font-size: 0.84rem; color: #0d3c6e; margin: 8px 0;
  direction: rtl;
}
.al-ok {
  background: #e8f5e9; border-right: 4px solid #388e3c;
  border-radius: 8px; padding: 10px 14px;
  font-size: 0.84rem; color: #1b5020; margin: 8px 0;
  direction: rtl;
}
.al-warn {
  background: #fff8e1; border-right: 4px solid #f9a825;
  border-radius: 8px; padding: 10px 14px;
  font-size: 0.84rem; color: #5d4300; margin: 8px 0;
  direction: rtl;
}

/* ── Buttons ─────────────────────────────────────── */
div.stButton > button {
  font-family: 'Cairo', sans-serif !important;
  font-weight: 700 !important;
}
div.stButton > button[kind="primary"] {
  background: linear-gradient(135deg, #0f0e0d, #2a1e08) !important;
  color: #b8933a !important;
  border: none !important;
}
div.stButton > button:hover { opacity: 0.88 !important; }

/* ── Gold Divider ────────────────────────────────── */
.gdiv {
  height: 1px; border: none; margin: 20px 0;
  background: linear-gradient(90deg, transparent, rgba(184,147,58,0.4), transparent);
}

/* ── Progress Item ───────────────────────────────── */
.prog-ok  { background:#e8f5e9; border-radius:8px; padding:6px 12px; margin:3px 0; font-size:0.82rem; color:#1b5020; }
.prog-err { background:#fdecea; border-radius:8px; padding:6px 12px; margin:3px 0; font-size:0.82rem; color:#b71c1c; }
.prog-run { background:#fff8e1; border-radius:8px; padding:6px 12px; margin:3px 0; font-size:0.82rem; color:#e65100; }

/* ── Badges ──────────────────────────────────────── */
.badge-ok   { display:inline-block; background:#e8f5e9; color:#2d7a4f; padding:2px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }
.badge-miss { display:inline-block; background:#fafafa; color:#9e9e9e; padding:2px 10px; border-radius:20px; font-size:0.72rem; }
.badge-new  { display:inline-block; background:#fff3e0; color:#e65100; padding:2px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }

/* ── Compare Card ────────────────────────────────── */
.cmp-card {
  background: white;
  border: 1px solid rgba(184,147,58,0.25);
  border-radius: 12px;
  padding: 14px;
  margin-bottom: 12px;
  direction: rtl;
}
.cmp-card.suspect { border-color: #f9a825; background: #fffde7; }
.cmp-card.exact   { border-color: #388e3c; background: #f1f8e9; }
.cmp-card img { width: 80px; height: 80px; object-fit: cover; border-radius: 8px; }
.cmp-pct { font-size: 1.2rem; font-weight: 900; color: #f9a825; }

/* ── Footer ──────────────────────────────────────── */
.mhw-footer {
  text-align: center; color: #9a8e80;
  font-size: 0.76rem; padding: 16px 0 8px;
  border-top: 1px solid rgba(184,147,58,0.15);
  margin-top: 32px;
}
</style>
""", unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  SALLA EXACT SCHEMAS                                            ║
# ╚══════════════════════════════════════════════════════════════════╝
SALLA_COLS = [
    "No.", "النوع ", "أسم المنتج", "تصنيف المنتج", "صورة المنتج",
    "وصف صورة المنتج", "نوع المنتج", "سعر المنتج", "الوصف",
    "هل يتطلب شحن؟", "رمز المنتج sku", "سعر التكلفة", "السعر المخفض",
    "تاريخ بداية التخفيض", "تاريخ نهاية التخفيض",
    "اقصي كمية لكل عميل", "إخفاء خيار تحديد الكمية",
    "اضافة صورة عند الطلب", "الوزن", "وحدة الوزن",
    "حالة المنتج", "الماركة", "العنوان الترويجي", "تثبيت المنتج",
    "الباركود", "السعرات الحرارية", "MPN", "GTIN",
    "خاضع للضريبة ؟", "سبب عدم الخضوع للضريبة",
    "[1] الاسم", "[1] النوع", "[1] القيمة", "[1] الصورة / اللون",
    "[2] الاسم", "[2] النوع", "[2] القيمة", "[2] الصورة / اللون",
    "[3] الاسم", "[3] النوع", "[3] القيمة", "[3] الصورة / اللون",
]

SALLA_SEO_COLS = [
    "No. (غير قابل للتعديل)",
    "اسم المنتج (غير قابل للتعديل)",
    "رابط مخصص للمنتج (SEO Page URL)",
    "عنوان صفحة المنتج (SEO Page Title)",
    "وصف صفحة المنتج (SEO Page Description)",
]

SALLA_PRICE_COLS = [
    "No.", "النوع ", "أسم المنتج", "رمز المنتج sku",
    "سعر المنتج", "سعر التكلفة", "السعر المخفض",
    "تاريخ بداية التخفيض", "تاريخ نهاية التخفيض",
]

# Salla brands file exact columns
SALLA_BRANDS_COLS = [
    "اسم العلامة التجارية",
    "(SEO Page URL) رابط صفحة العلامة التجارية",
    "وصف العلامة التجارية",
    "صورة العلامة التجارية",
]

# Editor shows these by default (rest hidden unless user selects)
EDITOR_COLS = [
    "No.", "النوع ", "أسم المنتج", "الماركة", "تصنيف المنتج",
    "سعر المنتج", "رمز المنتج sku", "صورة المنتج",
    "وصف صورة المنتج", "حالة المنتج", "السعر المخفض",
]

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

# ╔══════════════════════════════════════════════════════════════════╗
# ║  AI SYSTEM PROMPT — خبير وصف منتجات مهووس v4.5                ║
# ╚══════════════════════════════════════════════════════════════════╝
AI_SYSTEM = """أنت خبير كتابة أوصاف عطور فاخرة تعمل حصرياً لمتجر "مهووس" السعودي.

قواعد صارمة لا تُكسر:
- ممنوع منعاً باتاً استخدام الرموز التعبيرية (Emojis) نهائياً
- التركيز يُكتب دائماً: "أو دو بارفيوم"
- أسلوبك: راقٍ 40%، ودود 25%، رومانسي 20%، تسويقي مقنع 15%
- الطول: 1200-1500 كلمة بالضبط
- الإخراج HTML خالص فقط — لا نص خارج الوسوم
- استخدم <strong> للكلمات المفتاحية
- الروابط الداخلية: استخدم <a href="https://mahwous.com/brands/[slug]" target="_blank">[اسم الماركة]</a>
- المكونات: اذكر مكونات حقيقية موثوقة إذا عرفتها، وإلا اذكر مكونات تقريبية منطقية للعائلة العطرية

هيكل الوصف الإلزامي:
<h2>[عطر/تستر] [الماركة] [الاسم] [التركيز] [الحجم] [للجنس]</h2>
<p>فقرة افتتاحية عاطفية 100-150 كلمة، الكلمة المفتاحية في أول 50 كلمة، دعوة للشراء.</p>
<h3>تفاصيل المنتج</h3>
<ul>
<li><strong>الماركة:</strong> [مع رابط داخلي]</li>
<li><strong>الاسم:</strong></li>
<li><strong>الجنس:</strong></li>
<li><strong>العائلة العطرية:</strong></li>
<li><strong>الحجم:</strong></li>
<li><strong>التركيز:</strong> أو دو بارفيوم</li>
<li><strong>سنة الإصدار:</strong></li>
<li><strong>نوع المنتج:</strong> [تستر / عادي]</li>
</ul>
<h3>رحلة العطر - الهرم العطري</h3>
<p>وصف حسي شاعري للعطر كاملاً.</p>
<ul>
<li><strong>المقدمة (Top Notes):</strong> [المكونات الحقيقية أو التقريبية]</li>
<li><strong>القلب (Heart Notes):</strong> [المكونات الحقيقية أو التقريبية]</li>
<li><strong>القاعدة (Base Notes):</strong> [المكونات الحقيقية أو التقريبية]</li>
</ul>
<h3>لماذا تختار هذا العطر؟</h3>
<ul>
<li><strong>الثبات والفوحان:</strong> [وصف دقيق]</li>
<li><strong>التميز والأصالة:</strong> [وصف دقيق]</li>
<li><strong>القيمة الاستثنائية:</strong> [وصف دقيق]</li>
<li><strong>الجاذبية المضمونة:</strong> [وصف دقيق]</li>
</ul>
<h3>متى وأين ترتديه؟</h3>
<p>الفصول المناسبة، أوقات الاستخدام، المناسبات الملائمة.</p>
<h3>لمسة خبير من مهووس</h3>
<p>تقييم للفوحان (1-10) والثبات (1-10) ونصيحة رش احترافية.</p>
<h3>الأسئلة الشائعة</h3>
<ul>
<li><strong>كم يدوم العطر؟</strong> [إجابة دقيقة]</li>
<li><strong>هل يناسب الاستخدام اليومي؟</strong> [إجابة]</li>
<li><strong>ما الفرق بين التستر والعطر العادي؟</strong> [إجابة]</li>
<li><strong>ما العائلة العطرية؟</strong> [إجابة]</li>
<li><strong>هل يناسب الطقس الحار في السعودية؟</strong> [إجابة]</li>
<li><strong>ما مناسبات ارتداء هذا العطر؟</strong> [إجابة]</li>
</ul>
<h3>اكتشف أكثر من مهووس</h3>
<p>روابط داخلية لعطور مشابهة: <a href="https://mahwous.com/brands/[slug]" target="_blank">[عطور الماركة]</a> | <a href="https://mahwous.com/categories/mens-perfumes" target="_blank">[عطور رجالية]</a></p>
<p><strong>عالمك العطري يبدأ من مهووس.</strong> أصلي 100% | شحن سريع داخل السعودية.</p>

في نهاية الوصف أضف قسم SEO منفصل بصيغة JSON:
<!--SEO_DATA
{
  "page_title": "...",
  "meta_description": "...",
  "url_slug": "...",
  "alt_text": "...",
  "tags": ["...", "..."]
}
SEO_DATA-->"""

# ╔══════════════════════════════════════════════════════════════════╗
# ║  SESSION STATE INIT                                             ║
# ╚══════════════════════════════════════════════════════════════════╝
def _init_state():
    defaults = {
        # API keys
        "api_key":        os.environ.get("ANTHROPIC_API_KEY", ""),
        "google_api":     os.environ.get("GOOGLE_API_KEY", ""),
        "google_cse":     os.environ.get("GOOGLE_CSE_ID", ""),
        # Reference data
        "brands_df":      None,
        "categories_df":  None,
        # Universal Processor state
        "up_raw":         None,   # raw uploaded df
        "up_df":          None,   # restructured Salla df
        "up_seo":         None,   # SEO companion df
        "up_filename":    "",
        "up_mapped":      False,
        # Quick Add list
        "qa_rows":        [],
        # Comparison page state
        "cmp_new_df":     None,   # new products file
        "cmp_store_df":   None,   # store master file
        "cmp_results":    None,   # comparison results df
        "cmp_approved":   {},     # {idx: True/False} user decisions
        # New brands generated
        "new_brands":     [],     # list of dicts for new brands
        # Compare v9.4 page state
        "cv2_store_df":   None,   # ملف المتجر لصفحة المقارنة v9.4
        "cv2_comp_dfs":   [],     # ملفات المنافسين
        "cv2_brands_df":  None,   # ملف الماركات الخاص
        "cv2_results":    None,   # نتائج المقارنة
        "cv2_running":    False,
        "cv2_logs":       [],
        # Store Audit page state
        "audit_df":       None,   # ملف المتجر للتدقيق
        "audit_results":  None,   # نتائج التدقيق
        # Page
        "page":           "compare_v2",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # Auto-load bundled reference CSVs
    if st.session_state.brands_df is None:
        p = os.path.join(DATA_DIR, "brands.csv")
        if os.path.exists(p):
            try:
                st.session_state.brands_df = pd.read_csv(p, encoding="utf-8-sig")
            except Exception:
                pass
    if st.session_state.categories_df is None:
        p = os.path.join(DATA_DIR, "categories.csv")
        if os.path.exists(p):
            try:
                st.session_state.categories_df = pd.read_csv(p, encoding="utf-8-sig")
            except Exception:
                pass

_init_state()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  CORE UTILITIES                                                 ║
# ╚══════════════════════════════════════════════════════════════════╝

def read_file(f, salla_2row: bool = False) -> pd.DataFrame:
    """Read CSV or Excel → clean DataFrame. Handles multi-encoding."""
    name = f.name.lower()
    hdr  = 1 if salla_2row else 0
    try:
        if name.endswith((".xlsx", ".xlsm", ".xls")):
            df = pd.read_excel(f, header=hdr, dtype=str)
        else:
            for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
                try:
                    f.seek(0)
                    df = pd.read_csv(f, header=hdr, encoding=enc, dtype=str)
                    break
                except UnicodeDecodeError:
                    continue
        df = df.dropna(how="all").reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception as e:
        st.error(f"خطأ في قراءة الملف: {e}")
        return pd.DataFrame()


def auto_guess_col(cols, keywords: list) -> str:
    """Guess which column matches a list of keywords."""
    for kw in keywords:
        for c in cols:
            if kw.lower() in c.lower():
                return c
    return "— لا يوجد —"


def _fuzzy_ratio(a: str, b: str) -> int:
    """Similarity ratio (0-100) — uses rapidfuzz when available for higher accuracy."""
    a, b = str(a).lower().strip(), str(b).lower().strip()
    if not a or not b:
        return 0
    if a == b:
        return 100
    if HAS_RAPIDFUZZ:
        return int(rf_fuzz.token_set_ratio(a, b))
    # Fallback: LCS-based ratio
    longer  = max(len(a), len(b))
    matches = 0
    j = 0
    for ch in a:
        while j < len(b):
            if b[j] == ch:
                matches += 1
                j += 1
                break
            j += 1
    return int(matches / longer * 100)


# ══════════════════════════════════════════════════════════════════
#  خوارزميات المقارنة الذكية v9.4 — MahwousEngine
# ══════════════════════════════════════════════════════════════════

_CATEGORY_MAP_V94 = {
    "تستر":        "العطور > تستر",
    "طقم هدايا":   "العطور > طقم هدايا",
    "عطر شعر":     "العطور > عطور الشعر",
    "عناية جسم":   "العناية > لوشن وكريم",
    "شاور جل":     "العناية > شاور جل",
    "مزيل عرق":    "العناية > مزيل العرق",
    "معطر جسم":    "العطور > معطر جسم",
    "عطر تجاري":   "العطور",
}


def extract_product_attrs(name: str) -> dict:
    """استخراج الحجم، النوع، التركيز، والاسم النقي من اسم المنتج."""
    s = str(name).lower().strip()

    # الحجم
    m = re.search(r"(\d+)\s*(?:مل|ml|ملل|cc)", s, re.IGNORECASE)
    size = int(m.group(1)) if m else 0

    # النوع
    ptype = "عطر تجاري"
    if any(w in s for w in ["تستر", "tester", "بدون كرتون", "ديمو", "demo"]):
        ptype = "تستر"
    elif any(w in s for w in ["طقم", "مجموعة", "set ", "gift"]):
        ptype = "طقم هدايا"
    elif any(w in s for w in ["عطر شعر", "hair mist", "للشعر"]):
        ptype = "عطر شعر"
    elif any(w in s for w in ["لوشن", "lotion", "كريم", "cream", "body butter"]):
        ptype = "عناية جسم"
    elif any(w in s for w in ["شاور", "shower", "جل استحمام", "bath"]):
        ptype = "شاور جل"
    elif any(w in s for w in ["مزيل", "deodorant", "ديودرنت", "roll-on", "stick"]):
        ptype = "مزيل عرق"
    elif any(w in s for w in ["بدي مست", "body mist", "معطر جسم", "body spray"]):
        ptype = "معطر جسم"

    # التركيز
    conc = "غير محدد"
    if any(w in s for w in ["extrait", "اكستريت"]):
        conc = "Extrait"
    elif any(w in s for w in ["edp", "eau de parfum", "او دي بارفيوم", "او دو بارفيوم",
                               "بارفيوم", "de parfum", "برفيوم", "le parfum"]):
        conc = "EDP"
    elif any(w in s for w in ["edt", "eau de toilette", "او دي تواليت", "او دو تواليت",
                               "toilette", "تواليت"]):
        conc = "EDT"
    elif any(w in s for w in ["pure parfum", "parfum", "بارفان"]):
        conc = "Parfum"
    elif any(w in s for w in ["edc", "cologne", "كولونيا"]):
        conc = "EDC"
    if any(w in s for w in ["intense", "انتنس", "انتنز"]):
        conc += " Intense"
    if any(w in s for w in ["absolu", "ابسولو"]):
        conc += " Absolu"

    # الاسم النقي
    clean = re.sub(r"\d+\s*(?:مل|ml|ملل|cc|g|جرام|oz|x\s*\d+)", "", s)
    strip_words = [
        "eau de parfum", "eau de toilette", "le parfum", "de parfum",
        "او دي بارفيوم", "او دو بارفيوم", "او دي تواليت", "او دو تواليت",
        "edp", "edt", "edc", "extrait", "parfum", "perfume", "cologne",
        "بارفيوم", "برفيوم", "بارفان", "تواليت", "اكستريت", "كولونيا",
        "عطر", "طقم", "مجموعة", "تستر", "tester", "للرجال", "للنساء",
        "نسائي", "رجالي", "للجنسين", "مركز", "hair mist", "body mist",
        "شاور جل", "لوشن", "set", "intense", "absolu", "انتنس", "ابسولو",
        "للشعر", "spray", "بدون كرتون", "gift",
    ]
    for w in sorted(strip_words, key=len, reverse=True):
        clean = clean.replace(w, " ")
    clean = re.sub(r"\s+", " ", clean).strip()

    return {
        "size": size,
        "type": ptype,
        "concentration": conc,
        "clean_name": clean,
        "category": _CATEGORY_MAP_V94.get(ptype, "العطور"),
    }


def run_smart_comparison(new_df: pd.DataFrame, store_df: pd.DataFrame,
                          new_name_col: str, store_name_col: str,
                          new_sku_col: str = None, store_sku_col: str = None,
                          new_img_col: str = None,
                          t_dup: int = 88, t_near: int = 75, t_review: int = 55,
                          brands_list: list = None) -> pd.DataFrame:
    """
    خوارزمية المقارنة الذكية v9.4 المدمجة في Streamlit.
    تصنّف كل منتج جديد إلى: مكرر / مراجعة يدوية / فرصة جديدة
    مع استخدام rapidfuzz + تحليل الحجم والتركيز والنوع.
    """
    if brands_list is None:
        brands_list = []
    brands_lower = [b.lower() for b in brands_list]

    # تفكيك منتجات المتجر
    store_parsed = []
    for _, row in store_df.iterrows():
        sname = str(row.get(store_name_col, "") or "").strip()
        if not sname or sname == "nan":
            continue
        attrs = extract_product_attrs(sname)
        store_parsed.append({
            "orig_name": sname,
            "clean_name": attrs["clean_name"],
            "size": attrs["size"],
            "type": attrs["type"],
            "concentration": attrs["concentration"],
            "sku": str(row.get(store_sku_col, "") or "") if store_sku_col else "",
            "image": str(row.get("صورة المنتج", "") or ""),
            "price": str(row.get("سعر المنتج", "") or ""),
        })
    store_clean_dict = {i: p["clean_name"] for i, p in enumerate(store_parsed)}
    store_sku_set = {p["sku"].lower() for p in store_parsed if p["sku"]}

    results = []
    for i, row in new_df.iterrows():
        new_name = str(row.get(new_name_col, "") or "").strip()
        new_sku  = str(row.get(new_sku_col, "") or "").strip() if new_sku_col else ""
        new_img  = str(row.get(new_img_col, "") or "").strip() if new_img_col else                    str(row.get("صورة المنتج", "") or "").strip()
        if not new_name or new_name == "nan":
            continue

        # تفكيك المنتج الجديد
        new_attrs = extract_product_attrs(new_name)
        new_clean = new_attrs["clean_name"]
        new_size  = new_attrs["size"]
        new_type  = new_attrs["type"]
        new_conc  = new_attrs["concentration"]
        new_cat   = new_attrs["category"]

        # كشف الماركة
        brand = ""
        nl = new_name.lower()
        for b, bo in zip(brands_lower, brands_list):
            if b in nl:
                brand = bo
                break
        if not brand:
            words = new_name.split()
            brand = " ".join(words[:2]) if len(words) >= 2 else (words[0] if words else "")

        # تطابق SKU مباشر
        if new_sku and new_sku.lower() in store_sku_set:
            results.append({
                "الاسم الجديد":           new_name,
                "SKU الجديد":             new_sku,
                "الماركة":                brand,
                "التصنيف":                new_cat,
                "أقرب تطابق في المتجر":   new_name,
                "نسبة التشابه":           100,
                "الحالة":                 "مكرر (SKU)",
                "سبب القرار":             "تطابق SKU مباشر",
                "الإجراء":                "حذف",
                "_idx":                   i,
                "_img":                   new_img,
            })
            continue

        # مطابقة fuzzy ذكية
        verdict = "فرصة جديدة"
        reason  = "منتج جديد — لا يوجد تشابه مع متجرنا"
        score   = 0
        best_store_name = ""

        if store_clean_dict:
            if HAS_RAPIDFUZZ:
                best = rf_process.extractOne(new_clean, store_clean_dict,
                                             scorer=rf_fuzz.token_set_ratio)
                if best:
                    _, raw_score, pos = best
                    score = raw_score
                    sp = store_parsed[pos]
                    best_store_name = sp["orig_name"]
                    s_size = sp["size"]
                    s_type = sp["type"]
                    s_conc = sp["concentration"]
                else:
                    raw_score = 0
            else:
                # Fallback بدون rapidfuzz
                raw_score = 0
                pos = -1
                for idx2, clean2 in store_clean_dict.items():
                    s = _fuzzy_ratio(new_clean, clean2)
                    if s > raw_score:
                        raw_score = s
                        pos = idx2
                score = raw_score
                if pos >= 0:
                    sp = store_parsed[pos]
                    best_store_name = sp["orig_name"]
                    s_size = sp["size"]
                    s_type = sp["type"]
                    s_conc = sp["concentration"]

            if raw_score >= t_dup:
                if new_type != s_type:
                    verdict = "فرصة جديدة"
                    reason  = f"نوع مختلف — المنافس: ({new_type}) | متجرنا: ({s_type})"
                elif new_size != s_size and new_size != 0 and s_size != 0:
                    verdict = "فرصة جديدة"
                    reason  = f"حجم مختلف — المنافس: ({new_size}مل) | متجرنا: ({s_size}مل)"
                elif (new_conc != s_conc and new_conc != "غير محدد" and s_conc != "غير محدد"):
                    verdict = "فرصة جديدة"
                    reason  = f"تركيز مختلف — المنافس: ({new_conc}) | متجرنا: ({s_conc})"
                else:
                    verdict = "مكرر"
                    reason  = f"تطابق تام ({raw_score:.0f}%) — اسم + حجم + تركيز + نوع"
            elif raw_score >= t_near:
                if new_type != s_type:
                    verdict = "فرصة جديدة"
                    reason  = f"تشابه قوي ({raw_score:.0f}%) لكن النوع مختلف"
                elif new_size != s_size and new_size != 0 and s_size != 0:
                    verdict = "فرصة جديدة"
                    reason  = f"تشابه قوي ({raw_score:.0f}%) لكن الحجم مختلف"
                else:
                    verdict = "مراجعة يدوية"
                    reason  = f"تشابه ({raw_score:.0f}%) — راجع يدوياً"
            elif raw_score >= t_review:
                verdict = "مراجعة يدوية"
                reason  = f"تشابه جزئي ({raw_score:.0f}%) — راجع يدوياً"

        # تحويل الحكم إلى حالة/إجراء
        if verdict == "مكرر":
            status = "مكرر"
            action = "حذف"
        elif verdict == "مراجعة يدوية":
            status = "مشبوه"
            action = "مراجعة"
        else:
            status = "جديد"
            action = "اعتماد"

        results.append({
            "الاسم الجديد":           new_name,
            "SKU الجديد":             new_sku,
            "الماركة":                brand,
            "التصنيف":                new_cat,
            "أقرب تطابق في المتجر":   best_store_name,
            "نسبة التشابه":           score,
            "الحالة":                 status,
            "سبب القرار":             reason,
            "الإجراء":                action,
            "_idx":                   i,
            "_img":                   new_img,
        })

    return pd.DataFrame(results) if results else pd.DataFrame()


def match_brand(name: str) -> dict:
    bdf = st.session_state.brands_df
    if bdf is None or not str(name).strip():
        return {"name": "", "page_url": ""}
    nl = str(name).lower()
    col0 = bdf.columns[0]
    for _, row in bdf.iterrows():
        raw = str(row[col0])
        for part in re.split(r"\s*\|\s*", raw):
            p = part.strip().lower()
            if p and p in nl:
                return {
                    "name": raw,
                    "page_url": str(row.get(
                        "(SEO Page URL) رابط صفحة العلامة التجارية", "") or ""),
                }
    return {"name": "", "page_url": ""}


def generate_new_brand(brand_name: str) -> dict:
    """Generate a new brand entry in Salla format using AI."""
    key = st.session_state.api_key
    slug = to_slug(brand_name)
    desc = ""
    if key:
        try:
            client = anthropic.Anthropic(api_key=key)
            msg = client.messages.create(
                model="claude-opus-4-5",
                max_tokens=300,
                messages=[{"role": "user", "content":
                    f"اكتب وصفاً موجزاً (50-80 كلمة) لعلامة العطور التجارية '{brand_name}' "
                    f"بالعربية، بأسلوب فاخر ومهني. بدون رموز تعبيرية. نص فقط."}],
            )
            desc = msg.content[0].text.strip()
        except Exception:
            desc = f"علامة تجارية عالمية متخصصة في صناعة العطور الفاخرة."
    return {
        "اسم العلامة التجارية":                           brand_name,
        "(SEO Page URL) رابط صفحة العلامة التجارية":     slug,
        "وصف العلامة التجارية":                          desc,
        "صورة العلامة التجارية":                         "",
    }


def match_category(name: str, gender: str = "") -> str:
    t = (str(name) + " " + str(gender)).lower()
    if any(w in t for w in ["رجال", "للرجال", "men", "homme", "رجالي"]):
        return "العطور > عطور رجالية"
    if any(w in t for w in ["نساء", "للنساء", "women", "femme", "نسائي"]):
        return "العطور > عطور نسائية"
    return "العطور > عطور للجنسين"


def to_slug(text: str) -> str:
    ar = {
        "ا": "a", "أ": "a", "إ": "e", "آ": "a", "ب": "b", "ت": "t",
        "ث": "th", "ج": "j", "ح": "h", "خ": "kh", "د": "d", "ذ": "z",
        "ر": "r", "ز": "z", "س": "s", "ش": "sh", "ص": "s", "ض": "d",
        "ط": "t", "ظ": "z", "ع": "a", "غ": "gh", "ف": "f", "ق": "q",
        "ك": "k", "ل": "l", "م": "m", "ن": "n", "ه": "h", "و": "w",
        "ي": "y", "ى": "a", "ة": "a", "ء": "", "ئ": "y", "ؤ": "w",
    }
    out = ""
    for c in str(text).lower():
        if c in ar:
            out += ar[c]
        elif c.isascii() and c.isalnum():
            out += c
        elif c in " -_":
            out += "-"
    return re.sub(r"-+", "-", out).strip("-") or "perfume"


def gen_seo(name: str, brand: dict, size: str,
            tester: bool, gender: str) -> dict:
    bname = brand.get("name", "")
    parts = re.split(r"\s*\|\s*", bname)
    ben   = parts[-1].strip() if len(parts) > 1 else bname
    pref  = "تستر" if tester else "عطر"
    title = f"{pref} {name} {size} | {ben}".strip()
    desc  = (f"تسوق {pref} {name} {size} الأصلي من {bname}. "
             f"عطر {gender} فاخر ثابت. أصلي 100% من مهووس.")
    if len(desc) > 160:
        desc = desc[:157] + "..."
    slug = to_slug(f"{ben}-{name}-{size}".replace("مل", "ml"))
    return {
        "url":   slug,
        "title": title,
        "desc":  desc,
        "alt":   f"زجاجة {pref} {name} {size} الأصلية",
    }


def fetch_image(name: str, tester: bool = False) -> str:
    gk = st.session_state.google_api
    cx = st.session_state.google_cse
    if not gk or not cx:
        return ""
    try:
        q = name + (" tester box" if tester else " perfume bottle")
        r = requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params={"key": gk, "cx": cx, "q": q,
                    "searchType": "image", "num": 1, "imgSize": "large"},
            timeout=10,
        )
        items = r.json().get("items", [])
        return items[0]["link"] if items else ""
    except Exception:
        return ""


def scrape_product_url(url: str) -> dict:
    """سحب بيانات المنتج من رابط URL مع التعامل مع Cloudflare وغيره.
    يعيد dict يحتوي على: name, price, image, images, desc, brand_hint
    """
    result = {"name": "", "price": "", "image": "", "images": [], "desc": "", "brand_hint": "", "error": ""}
    try:
        from bs4 import BeautifulSoup
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "ar,en-US;q=0.9,en;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
        }
        resp = requests.get(url, headers=headers, timeout=20, allow_redirects=True)
        if resp.status_code != 200:
            result["error"] = f"خطأ HTTP {resp.status_code}"
            return result
        soup = BeautifulSoup(resp.text, "html.parser")

        # ── Extract Name ──────────────────────────────────────────────
        name = ""
        # Try og:title first
        og_title = soup.find("meta", property="og:title")
        if og_title and og_title.get("content"):
            name = og_title["content"].strip()
        if not name:
            h1 = soup.find("h1")
            if h1:
                name = h1.get_text(" ", strip=True)
        if not name:
            title_tag = soup.find("title")
            if title_tag:
                name = title_tag.get_text(strip=True).split("|")[0].split("-")[0].strip()
        result["name"] = name[:200] if name else ""

        # ── Extract Price ─────────────────────────────────────────────
        price = ""
        # Try structured data (JSON-LD)
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string or "{}")
                if isinstance(data, list):
                    data = data[0]
                offers = data.get("offers", data.get("Offers", {}))
                if isinstance(offers, list):
                    offers = offers[0]
                p = offers.get("price", "")
                if p:
                    price = str(p)
                    break
            except Exception:
                pass
        if not price:
            # Try og:price
            og_price = soup.find("meta", property="product:price:amount")
            if og_price and og_price.get("content"):
                price = og_price["content"]
        if not price:
            # Try common price selectors
            for sel in [".price", "[class*='price']", "[itemprop='price']", ".product-price"]:
                el = soup.select_one(sel)
                if el:
                    txt = el.get_text(strip=True)
                    nums = re.findall(r'[\d,\.]+', txt)
                    if nums:
                        price = nums[0].replace(",", "")
                        break
        result["price"] = price

        # ── Extract Images ────────────────────────────────────────────
        images = []
        og_img = soup.find("meta", property="og:image")
        if og_img and og_img.get("content"):
            images.append(og_img["content"])
        # Try JSON-LD images
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string or "{}")
                if isinstance(data, list):
                    data = data[0]
                imgs = data.get("image", [])
                if isinstance(imgs, str):
                    imgs = [imgs]
                for img in imgs[:5]:
                    if img and img not in images:
                        images.append(img)
            except Exception:
                pass
        # Try product gallery images
        for img_tag in soup.select("img[src]"):
            src = img_tag.get("src", "")
            if any(kw in src.lower() for kw in ["product", "item", "shop", "catalog", "perfume", "bottle"]):
                if src.startswith("http") and src not in images:
                    images.append(src)
                    if len(images) >= 6:
                        break
        result["image"] = images[0] if images else ""
        result["images"] = images[:6]

        # ── Extract Description ───────────────────────────────────────
        desc = ""
        og_desc = soup.find("meta", property="og:description")
        if og_desc and og_desc.get("content"):
            desc = og_desc["content"].strip()
        if not desc:
            meta_desc = soup.find("meta", attrs={"name": "description"})
            if meta_desc and meta_desc.get("content"):
                desc = meta_desc["content"].strip()
        if not desc:
            for sel in [".product-description", "[class*='description']", "[itemprop='description']"]:
                el = soup.select_one(sel)
                if el:
                    desc = el.get_text(" ", strip=True)[:500]
                    break
        result["desc"] = desc[:500] if desc else ""

        # ── Extract Brand Hint ────────────────────────────────────────
        brand_hint = ""
        og_brand = soup.find("meta", property="product:brand")
        if og_brand and og_brand.get("content"):
            brand_hint = og_brand["content"]
        if not brand_hint:
            for script in soup.find_all("script", type="application/ld+json"):
                try:
                    data = json.loads(script.string or "{}")
                    if isinstance(data, list):
                        data = data[0]
                    b = data.get("brand", {})
                    if isinstance(b, dict):
                        brand_hint = b.get("name", "")
                    elif isinstance(b, str):
                        brand_hint = b
                    if brand_hint:
                        break
                except Exception:
                    pass
        result["brand_hint"] = brand_hint

    except requests.exceptions.Timeout:
        result["error"] = "انتهت مهلة الاتصال (timeout)"
    except requests.exceptions.ConnectionError:
        result["error"] = "تعذّر الاتصال بالموقع"
    except Exception as e:
        result["error"] = f"خطأ: {str(e)[:100]}"
    return result


def ai_generate(name: str, tester: bool, brand: dict,
                size: str, gender: str, conc: str) -> str:
    key = st.session_state.api_key
    if not key:
        return "<p>أضف مفتاح Anthropic API في الإعدادات أولاً</p>"
    try:
        client = anthropic.Anthropic(api_key=key)
        ptype  = "تستر" if tester else "عطر"
        blink  = ""
        if brand.get("page_url"):
            blink = (f'— <a href="https://mahwous.com/{brand["page_url"]}"'
                     f' target="_blank">{brand["name"]}</a>')
        msg = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=4096,
            system=AI_SYSTEM,
            messages=[{"role": "user", "content":
                f"اكتب وصفاً HTML احترافياً كاملاً:\n"
                f"- النوع: {ptype}\n"
                f"- الاسم: {name}\n"
                f"- الماركة: {brand.get('name', 'غير محدد')} {blink}\n"
                f"- الحجم: {size}\n"
                f"- التركيز: {conc}\n"
                f"- الجنس: {gender}\n"
                f"أعد HTML خالصاً فقط بدون أي نص خارجه."}],
        )
        return msg.content[0].text
    except Exception as e:
        return f"<p>خطأ في الذكاء الاصطناعي: {e}</p>"


def build_empty_salla_row() -> dict:
    r = {c: "" for c in SALLA_COLS}
    r["النوع "]                    = "منتج"
    r["نوع المنتج"]               = "منتج جاهز"
    r["هل يتطلب شحن؟"]           = "نعم"
    r["خاضع للضريبة ؟"]          = "نعم"
    r["الوزن"]                    = "0.2"
    r["وحدة الوزن"]               = "kg"
    r["حالة المنتج"]              = "مرئي"
    r["اقصي كمية لكل عميل"]      = "0"
    r["إخفاء خيار تحديد الكمية"] = "0"
    r["اضافة صورة عند الطلب"]    = "0"
    return r


def fill_row(name, price="", sku="", image="", desc="",
             brand=None, category="", seo=None, no="",
             weight="0.2", weight_unit="kg", size="") -> dict:
    if brand is None:
        brand = {}
    if seo is None:
        seo = {}
    r = build_empty_salla_row()
    r["No."]             = str(no)
    r["أسم المنتج"]      = str(name)
    r["سعر المنتج"]      = str(price)
    r["رمز المنتج sku"]  = str(sku)
    r["صورة المنتج"]     = str(image)
    r["وصف صورة المنتج"] = seo.get("alt", "")
    r["الوصف"]           = str(desc)
    r["الماركة"]         = brand.get("name", "")
    r["تصنيف المنتج"]    = str(category)
    r["الوزن"]           = str(weight) if weight else "0.2"
    r["وحدة الوزن"]      = str(weight_unit) if weight_unit else "kg"
    # If no price → set quantity to 0
    if not str(price).strip() or str(price).strip() in ("0", "nan", "None"):
        r["اقصي كمية لكل عميل"] = "0"
    return r

# ╔══════════════════════════════════════════════════════════════════╗
# ║  EXPORT FUNCTIONS                                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def _style_header_row(ws, row_num: int, cols: list,
                      bg: str = "0F0E0D", fg: str = "B8933A"):
    for i, col in enumerate(cols, 1):
        c = ws.cell(row_num, i, col)
        c.font      = Font(bold=True, color="FFFFFF" if bg != "E8D5B7" else "0F0E0D",
                           name="Cairo", size=9)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, readingOrder=2)
        c.border    = Border(bottom=Side(style="thin", color=fg))
    ws.row_dimensions[row_num].height = 30


def export_product_xlsx(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Salla Products Template Sheet"

    # Row 1 — merged section header
    ws.cell(1, 1, "بيانات المنتج")
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(SALLA_COLS))
    c = ws.cell(1, 1)
    c.font      = Font(bold=True, color="FFFFFF", name="Cairo", size=12)
    c.fill      = PatternFill("solid", fgColor="0F0E0D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2 — column names
    _style_header_row(ws, 2, SALLA_COLS, bg="E8D5B7", fg="B8933A")

    # Data rows from row 3
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(SALLA_COLS, 1):
            v = str(row.get(col, "") if pd.notna(row.get(col, "")) else "")
            c = ws.cell(ri, ci, v)
            c.alignment = Alignment(horizontal="right", vertical="top",
                                    wrap_text=(col == "الوصف"),
                                    readingOrder=2)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FAFAF8")
        ws.row_dimensions[ri].height = 18

    # Column widths
    W = {
        "أسم المنتج": 45, "الوصف": 55, "تصنيف المنتج": 38,
        "صورة المنتج": 46, "الماركة": 24, "No.": 13,
        "وصف صورة المنتج": 36,
    }
    for i, col in enumerate(SALLA_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = W.get(col, 14)
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_product_csv(df: pd.DataFrame) -> bytes:
    out = io.StringIO()
    # Row 1
    out.write("بيانات المنتج" + "," * (len(SALLA_COLS) - 1) + "\n")
    # Row 2
    out.write(",".join(SALLA_COLS) + "\n")
    for _, row in df.iterrows():
        vals = []
        for c in SALLA_COLS:
            v = str(row.get(c, "") if pd.notna(row.get(c, "")) else "")
            if any(x in v for x in [",", "\n", '"']):
                v = f'"{v.replace(chr(34), chr(34)*2)}"'
            vals.append(v)
        out.write(",".join(vals) + "\n")
    return out.getvalue().encode("utf-8-sig")


def export_seo_xlsx(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Salla Product Seo Sheet"
    _style_header_row(ws, 1, SALLA_SEO_COLS, bg="1A1510", fg="B8933A")
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(SALLA_SEO_COLS, 1):
            v = str(row.get(col, "") if pd.notna(row.get(col, "")) else "")
            c = ws.cell(ri, ci, v)
            c.alignment = Alignment(horizontal="right", vertical="top",
                                    wrap_text=True, readingOrder=2)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FFF8E1")
        ws.row_dimensions[ri].height = 18
    W2 = {"اسم المنتج (غير قابل للتعديل)": 50,
          "وصف صفحة المنتج (SEO Page Description)": 65,
          "عنوان صفحة المنتج (SEO Page Title)": 52}
    for i, col in enumerate(SALLA_SEO_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = W2.get(col, 22)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def export_seo_csv(df: pd.DataFrame) -> bytes:
    out = io.StringIO()
    out.write(",".join(SALLA_SEO_COLS) + "\n")
    for _, row in df.iterrows():
        vals = []
        for c in SALLA_SEO_COLS:
            v = str(row.get(c, "") if pd.notna(row.get(c, "")) else "")
            if any(x in v for x in [",", "\n"]):
                v = f'"{v}"'
            vals.append(v)
        out.write(",".join(vals) + "\n")
    return out.getvalue().encode("utf-8-sig")


def export_price_xlsx(df: pd.DataFrame) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Price Update"
    ws.cell(1, 1, "بيانات المنتج")
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(SALLA_PRICE_COLS))
    c = ws.cell(1, 1)
    c.font      = Font(bold=True, color="FFFFFF", name="Cairo")
    c.fill      = PatternFill("solid", fgColor="0F0E0D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    _style_header_row(ws, 2, SALLA_PRICE_COLS, bg="E8D5B7", fg="B8933A")
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(SALLA_PRICE_COLS, 1):
            ws.cell(ri, ci, str(row.get(col, "") or ""))
        ws.row_dimensions[ri].height = 18
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def export_price_csv(df: pd.DataFrame) -> bytes:
    out = io.StringIO()
    out.write("بيانات المنتج" + "," * (len(SALLA_PRICE_COLS) - 1) + "\n")
    out.write(",".join(SALLA_PRICE_COLS) + "\n")
    for _, row in df.iterrows():
        out.write(",".join([f'"{str(row.get(c,"") or "")}"'
                            for c in SALLA_PRICE_COLS]) + "\n")
    return out.getvalue().encode("utf-8-sig")


def export_brands_xlsx(brands_list: list) -> bytes:
    """Export new brands in Salla brands file format."""
    wb = Workbook(); ws = wb.active; ws.title = "New Brands"
    _style_header_row(ws, 1, SALLA_BRANDS_COLS, bg="0F0E0D", fg="B8933A")
    for ri, brand in enumerate(brands_list, 2):
        for ci, col in enumerate(SALLA_BRANDS_COLS, 1):
            ws.cell(ri, ci, str(brand.get(col, "") or ""))
        ws.row_dimensions[ri].height = 18
    for i, col in enumerate(SALLA_BRANDS_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = 40 if i == 3 else 28
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  SIDEBAR NAVIGATION                                             ║
# ╚══════════════════════════════════════════════════════════════════╝
with st.sidebar:
    st.markdown("""
    <div style="text-align:center;padding:18px 0 10px">
      <div style="font-size:2.4rem">🌸</div>
      <div style="color:#b8933a;font-size:1.25rem;font-weight:900;margin:4px 0">مهووس</div>
      <div style="color:rgba(255,255,255,0.3);font-size:0.7rem">مركز التحكم الشامل v4.8</div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    PAGES = [
        ("🔎", "مقارنة المنافسين",       "compare_v2"),
        ("🛠️", "المُعالج الشامل",       "processor"),
        ("🔀", "المقارنة والتدقيق",     "compare"),
        ("🏪", "مدقق ملف المتجر",       "store_audit"),
        ("➕", "منتج سريع",              "quickadd"),
        ("⚙️", "الإعدادات",             "settings"),
    ]
    for icon, label, key in PAGES:
        active = st.session_state.page == key
        if st.button(f"{icon}  {label}", use_container_width=True,
                     type="primary" if active else "secondary",
                     key=f"nav_{key}"):
            st.session_state.page = key
            st.rerun()

    st.divider()
    # Status
    bok = st.session_state.brands_df is not None
    cok = st.session_state.categories_df is not None
    aok = bool(st.session_state.api_key)
    gok = bool(st.session_state.google_api and st.session_state.google_cse)

    status_html = "".join([
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if bok else "❌"} '
        f'الماركات: {len(st.session_state.brands_df) if bok else "غير محملة"}</div>',
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if cok else "❌"} '
        f'التصنيفات: {len(st.session_state.categories_df) if cok else "غير محملة"}</div>',
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if aok else "❌"} '
        f'Claude API: {"متصل" if aok else "غير مضبوط"}</div>',
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if gok else "—"} '
        f'Google CSE: {"متصل" if gok else "غير مضبوط"}</div>',
    ])
    st.markdown(status_html, unsafe_allow_html=True)

    # New brands indicator
    if st.session_state.new_brands:
        st.divider()
        nb = len(st.session_state.new_brands)
        st.markdown(f'<div style="font-size:0.77rem;padding:3px 0;color:#f9a825">🆕 {nb} ماركة جديدة بانتظار التصدير</div>',
                    unsafe_allow_html=True)

    # Active file info
    if st.session_state.up_df is not None:
        st.divider()
        fname = st.session_state.up_filename
        nrows = len(st.session_state.up_df)
        st.markdown(f"""
        <div style="background:rgba(184,147,58,0.1);border-radius:8px;padding:10px;
                    font-size:0.78rem;border:1px solid rgba(184,147,58,0.25)">
          <div style="font-weight:800;margin-bottom:4px">📄 الملف النشط</div>
          <div style="color:#b8933a">{fname[:30]}</div>
          <div style="color:rgba(255,255,255,0.4)">{nrows} صف</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🗑️ إغلاق الملف", use_container_width=True):
            st.session_state.up_raw     = None
            st.session_state.up_df      = None
            st.session_state.up_seo     = None
            st.session_state.up_filename = ""
            st.session_state.up_mapped  = False
            st.rerun()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE HEADER                                                    ║
# ╚══════════════════════════════════════════════════════════════════╝
TITLES = {
    "compare_v2":  ("🔎 مقارنة المنافسين",      "ارفع ملف متجرنا + ملفات المنافسين — استخرج الفرص الجديدة بخوارزمية v9.4"),
    "processor":   ("🛠️ المُعالج الشامل",       "ارفع ملف منتجات خام — أكمل الوزن، الحجم، الماركة، الوصف، والقسم — صدّر لسلة"),
    "compare":     ("🔀 المقارنة والتدقيق",     "قارن المنتجات الجديدة بالمتجر — استبعد المكرر — اعتمد أو ألغِ المشبوه"),
    "store_audit": ("🏪 مدقق ملف المتجر",       "افحص ملف المتجر — اكتشف المنتجات الناقصة — عالجها — صدّر بتنسيق سلة"),
    "quickadd":    ("➕ منتج سريع",              "أدخل رابط منتج أو ارفع صورة وسيكمل النظام الباقي"),
    "brands":      ("🔍 مدقق الماركات",         "قارن قائمة ماركات بقاعدة بيانات مهووس"),
    "settings":    ("⚙️ الإعدادات",             "مفاتيح API وقواعد البيانات المرجعية"),
}
ttl, sub = TITLES.get(st.session_state.page, ("مهووس", ""))
st.markdown(f"""
<div class="mhw-header">
  <div class="emblem">م</div>
  <div><h1>{ttl}</h1><p>{sub}</p></div>
</div>
""", unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 2 — UNIVERSAL PROCESSOR                                   ║
# ╚══════════════════════════════════════════════════════════════════╝
if st.session_state.page == "processor":

    # ── STEP A & B: Upload & Auto-Process ─────────────────────────
    st.markdown("""<div class="sec-title"><div class="bar"></div><h3>الخطوة 1 — رفع وتجهيز الملف تلقائياً</h3></div>""",
                unsafe_allow_html=True)

    st.markdown("""<div class="al-info">
    ارفع ملف منتجات سلة (بصيغة CSV أو Excel). سيقوم النظام تلقائياً بالتعرف على الأعمدة، 
    استكمال الماركات الناقصة، توليد الأوصاف، تحديد القسم، وإضافة الوزن والحجم.
    </div>""", unsafe_allow_html=True)

    up_file = st.file_uploader(
        "ارفع ملف المنتجات هنا",
        type=["csv", "xlsx", "xls", "xlsm"],
        label_visibility="collapsed",
        key="proc_uploader",
    )

    if up_file:
        # Automatically assume it's a Salla file (2 rows header) for processing
        df_raw = read_file(up_file, salla_2row=True)
        
        if not df_raw.empty:
            st.session_state.up_raw = df_raw
            st.session_state.up_filename = up_file.name
            
            with st.spinner("جاري تحليل وتجهيز الملف بالكامل..."):
                rows_out = []
                seo_out  = []
                new_brands_found = []
                
                # Progress bar for processing
                prog_bar = st.progress(0)
                prog_text = st.empty()
                
                total_rows = len(df_raw)
                
                # Detect columns (fuzzy matching just in case)
                cols = list(df_raw.columns)
                col_name = auto_guess_col(cols, ["اسم","name","أسم المنتج"])
                col_price = auto_guess_col(cols, ["سعر","price"])
                col_sku = auto_guess_col(cols, ["sku","رمز"])
                col_img = auto_guess_col(cols, ["صورة","image"])
                col_desc = auto_guess_col(cols, ["وصف","desc"])
                col_brand = auto_guess_col(cols, ["ماركة","brand"])
                
                for idx, src in df_raw.iterrows():
                    prog_bar.progress(int((idx + 1) / total_rows * 100))
                    
                    def gv(col):
                        if col == "— لا يوجد —" or col not in df_raw.columns:
                            return ""
                        return str(src.get(col, "") or "").strip()

                    name = gv(col_name)
                    if not name or name.lower() in ("nan", "none", ""):
                        continue
                        
                    prog_text.text(f"جاري معالجة: {name[:40]}...")

                    price     = gv(col_price)
                    sku       = gv(col_sku)
                    img       = gv(col_img)
                    desc      = gv(col_desc)
                    brand_raw = gv(col_brand)
                    
                    is_test = "تستر" in name.lower() or "tester" in name.lower()
                    
                    # Auto-extract size, conc, gender from name
                    size_m = re.search(r"\d+\s*(?:مل|ml)", name, re.I)
                    size   = size_m.group() if size_m else "100 مل"
                    
                    gender = "للجنسين"
                    if any(w in name.lower() for w in ["نسائ","women","للنساء"]): gender = "للنساء"
                    elif any(w in name.lower() for w in ["رجال","men","للرجال"]): gender = "للرجال"
                    
                    conc = "أو دو بارفيوم"
                    if "تواليت" in name.lower() or "toilette" in name.lower(): conc = "أو دو تواليت"
                    elif "كولون" in name.lower() or "cologne" in name.lower(): conc = "أو دو كولون"
                    elif "بارفيوم" in name.lower() or "parfum" in name.lower(): conc = "بارفيوم"

                    # Brand logic
                    brand = match_brand(name)
                    if not brand.get("name"):
                        if brand_raw and brand_raw != 'nan':
                            b_name = brand_raw
                        else:
                            # Guess brand from first two words
                            words = name.split()
                            b_name = " ".join(words[:2]) if len(words) >= 2 else name
                            
                        brand = {"name": b_name, "page_url": to_slug(b_name)}
                        
                        existing_new = [b["اسم العلامة التجارية"] for b in st.session_state.new_brands]
                        if b_name not in existing_new and b_name not in new_brands_found:
                            new_brands_found.append(b_name)

                    cat = match_category(name, gender)
                    if is_test:
                        cat = "العطور > تستر"
                        
                    seo = gen_seo(name, brand, size, is_test, gender)
                    
                    # Generate Mock AI Description if empty
                    if not desc or desc == 'nan':
                        desc = f"""<div style="direction:rtl; text-align:right;">
  <h2>{name}</h2>
  <p>عطر فاخر من دار <strong>{brand['name']}</strong>. يتميز بتركيبة فريدة وثبات عالي.</p>
  <ul>
    <li><strong>الماركة:</strong> {brand['name']}</li>
    <li><strong>النوع:</strong> {'تستر' if is_test else 'عطر'}</li>
    <li><strong>الحجم:</strong> {size}</li>
    <li><strong>التركيز:</strong> {conc}</li>
    <li><strong>الجنس:</strong> {gender}</li>
  </ul>
  <p><em>احصل عليه الآن من مهووس العطور بأفضل الأسعار.</em></p>
</div>"""

                    nr = fill_row(name=name, price=price, sku=sku, image=img,
                                  desc=desc, brand=brand, category=cat, seo=seo,
                                  weight="0.2")
                    rows_out.append(nr)
                    seo_out.append({
                        "No. (غير قابل للتعديل)":            nr["No."],
                        "اسم المنتج (غير قابل للتعديل)":     name,
                        "رابط مخصص للمنتج (SEO Page URL)":   seo["url"],
                        "عنوان صفحة المنتج (SEO Page Title)": seo["title"],
                        "وصف صفحة المنتج (SEO Page Description)": seo["desc"],
                    })

                st.session_state.up_df     = pd.DataFrame(rows_out)
                st.session_state.up_seo    = pd.DataFrame(seo_out)
                st.session_state.up_mapped = True

                # Add new brands to session
                if new_brands_found:
                    for bn in new_brands_found:
                        st.session_state.new_brands.append({
                            "اسم العلامة التجارية": bn,
                            "(SEO Page URL) رابط صفحة العلامة التجارية": to_slug(bn),
                            "وصف العلامة التجارية": f"علامة تجارية متخصصة في العطور الفاخرة - {bn}",
                            "صورة العلامة التجارية": "",
                        })

                prog_text.empty()
                st.success(f"✅ تم تجهيز {len(rows_out)} منتج بالكامل بنجاح!")
                if new_brands_found:
                    st.info(f"🆕 تم استخراج {len(new_brands_found)} ماركة جديدة.")
                
                # Auto-rerun to show toolbox
                time.sleep(1)
                st.rerun()

    # ── STEP C: Toolbox + Editor ──────────────────────────────────
    if st.session_state.up_df is not None and st.session_state.up_mapped:
        df = st.session_state.up_df

        # Stats
        def _cnt(col):
            return int((df.get(col, pd.Series(dtype=str)).fillna("")
                        .str.strip() != "").sum())

        no_img   = int((df.get("صورة المنتج",   pd.Series(dtype=str)).fillna("").str.strip() == "").sum())
        no_desc  = int((df.get("الوصف",          pd.Series(dtype=str)).fillna("").str.strip() == "").sum())
        no_brand = int((df.get("الماركة",        pd.Series(dtype=str)).fillna("").str.strip() == "").sum())
        no_price = int((df.get("سعر المنتج",     pd.Series(dtype=str)).fillna("").str.strip().isin(["","0","nan"]).sum()))

        st.markdown(f"""
        <div class="stats-bar">
          <div class="stat-box"><div class="n">{len(df)}</div><div class="lb">إجمالي المنتجات</div></div>
          <div class="stat-box"><div class="n" style="color:{'#e53935' if no_img else '#43a047'}">{no_img}</div><div class="lb">بدون صورة</div></div>
          <div class="stat-box"><div class="n" style="color:{'#e53935' if no_desc else '#43a047'}">{no_desc}</div><div class="lb">بدون وصف</div></div>
          <div class="stat-box"><div class="n" style="color:{'#f9a825' if no_brand else '#43a047'}">{no_brand}</div><div class="lb">بدون ماركة</div></div>
          <div class="stat-box"><div class="n" style="color:{'#e53935' if no_price else '#43a047'}">{no_price}</div><div class="lb">بدون سعر</div></div>
        </div>
        """, unsafe_allow_html=True)

        # ── TOOLBOX TABS ──────────────────────────────────────────
        st.markdown("""<div class="sec-title"><div class="bar"></div>
        <h3>أدوات المعالجة الذكية</h3></div>""", unsafe_allow_html=True)

        tabs = st.tabs(["🤖 توليد الأوصاف", "🖼 جلب الصور",
                        "🏷 الماركات والتصنيفات", "➕ إضافة منتج", "⚡ عمليات مجمعة"])

        # ── Tab 0: AI Descriptions ─────────────────────────────
        with tabs[0]:
            st.markdown("**توليد الوصف الاحترافي بالذكاء الاصطناعي (Claude)**")
            gen_scope = st.radio("نطاق التوليد:", [
                "الصفوف التي ليس لها وصف فقط",
                "صف واحد بتحديده",
                "كل الصفوف (سيستغرق وقتاً)",
            ], horizontal=True, key="gen_scope")

            dft_conc_ai = st.selectbox("التركيز الافتراضي:", ["أو دو بارفيوم","أو دو كولون","أو دو تواليت","بارفيوم"], key="ai_conc")
            dft_gn_ai   = st.selectbox("الجنس الافتراضي:",  ["للجنسين","للرجال","للنساء"], key="ai_gn")
            dft_sz_ai   = st.text_input("الحجم الافتراضي:", "100 مل", key="ai_sz")

            if gen_scope == "صف واحد بتحديده":
                sel_ai = st.number_input("رقم الصف:", 0, max(0, len(df)-1), 0, key="ai_row")

            if st.button("✨ توليد الأوصاف الآن", type="primary", key="gen_desc_btn"):
                if not st.session_state.api_key:
                    st.error("أضف مفتاح Anthropic API في الإعدادات أولاً")
                else:
                    if gen_scope == "الصفوف التي ليس لها وصف فقط":
                        idxs = [i for i in range(len(df))
                                if not str(df.iloc[i].get("الوصف","")).strip()]
                    elif gen_scope == "صف واحد بتحديده":
                        idxs = [sel_ai]
                    else:
                        idxs = list(range(len(df)))

                    prog = st.progress(0); stat = st.empty()
                    for n, i in enumerate(idxs):
                        row  = df.iloc[i]
                        name = str(row.get("أسم المنتج","")).strip()
                        if not name: continue
                        stat.markdown(f'<div class="prog-run">توليد ({n+1}/{len(idxs)}): {name}</div>',
                                      unsafe_allow_html=True)
                        is_t  = any(w in name.lower() for w in ["تستر","tester"])
                        brand = {"name": str(row.get("الماركة","") or ""),
                                 "page_url": to_slug(str(row.get("الماركة","") or ""))}
                        size_m = re.search(r"\d+\s*(?:مل|ml)", name, re.I)
                        size   = size_m.group() if size_m else dft_sz_ai
                        gender = (str(row.get("تصنيف المنتج","")) + " " + name)
                        gender = ("للنساء" if any(w in gender for w in ["نسائ","women"])
                                  else "للرجال" if any(w in gender for w in ["رجال","men"])
                                  else dft_gn_ai)
                        desc = ai_generate(name, is_t, brand, size, gender, dft_conc_ai)
                        df.at[df.index[i], "الوصف"] = desc
                        prog.progress(int((n+1)/len(idxs)*100))

                    st.session_state.up_df = df
                    stat.markdown(f'<div class="prog-ok">✅ تم توليد {len(idxs)} وصف!</div>',
                                  unsafe_allow_html=True)
                    st.rerun()

        # ── Tab 1: Images ──────────────────────────────────────
        with tabs[1]:
            st.markdown("**جلب الصور تلقائياً عبر Google Custom Search**")
            if not (st.session_state.google_api and st.session_state.google_cse):
                st.markdown("""<div class="al-warn">
                أضف GOOGLE_API_KEY و GOOGLE_CSE_ID في الإعدادات لتفعيل جلب الصور.
                </div>""", unsafe_allow_html=True)

            img_scope = st.radio("نطاق الجلب:", [
                "الصفوف بدون صورة فقط",
                "كل الصفوف",
            ], horizontal=True, key="img_scope")
            add_test_kw = st.checkbox("إضافة كلمة 'tester box' للتستر", value=True, key="add_tk")

            if st.button("🖼 جلب الصور الآن", type="primary", key="fetch_img_btn"):
                idxs = ([i for i in range(len(df))
                         if not str(df.iloc[i].get("صورة المنتج","")).strip()]
                        if img_scope.startswith("الصفوف") else list(range(len(df))))
                prog = st.progress(0); stat = st.empty(); fetched = 0
                for n, i in enumerate(idxs):
                    name = str(df.iloc[i].get("أسم المنتج","")).strip()
                    if not name: continue
                    stat.markdown(
                        f'<div class="prog-run">🖼 ({n+1}/{len(idxs)}) {name}</div>',
                        unsafe_allow_html=True)
                    is_t = add_test_kw and any(w in name.lower() for w in ["تستر","tester"])
                    url  = fetch_image(name, is_t)
                    if url:
                        df.at[df.index[i], "صورة المنتج"] = url
                        fetched += 1
                    prog.progress(int((n+1)/len(idxs)*100))

                st.session_state.up_df = df
                stat.markdown(f'<div class="prog-ok">✅ تم جلب {fetched} صورة من {len(idxs)} صف</div>',
                              unsafe_allow_html=True)
                st.rerun()

            st.divider()
            st.markdown("**إضافة رابط صورة يدوياً لصف محدد:**")
            mi1, mi2, mi3 = st.columns([1, 4, 1])
            with mi1: man_row = st.number_input("رقم الصف", 0, max(0, len(df)-1), 0, key="man_r")
            with mi2: man_url = st.text_input("رابط الصورة", placeholder="https://...", key="man_u")
            with mi3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("حفظ", key="save_man"):
                    if man_url.startswith("http"):
                        df.at[df.index[man_row], "صورة المنتج"] = man_url
                        st.session_state.up_df = df
                        st.success("✅ تم حفظ الصورة")
                        st.rerun()

        # ── Tab 2: Brands & Categories ─────────────────────────
        with tabs[2]:
            st.markdown("**تعيين الماركات والتصنيفات التلقائي والمنابي**")
            scope_b = st.radio("نطاق:", [
                "الصفوف التي ليس لها ماركة فقط",
                "كل الصفوف (يُعيد التعيين)",
            ], horizontal=True, key="scope_b")

            if st.button("🏷 تعيين الآن", type="primary", key="assign_b"):
                idxs = ([i for i in range(len(df))
                         if not str(df.iloc[i].get("الماركة","")).strip()]
                        if scope_b.startswith("الصفوف") else list(range(len(df))))
                new_brands_auto = []
                for i in idxs:
                    name  = str(df.iloc[i].get("أسم المنتج","")).strip()
                    if not name: continue
                    brand = match_brand(name)
                    cat   = match_category(name)
                    if brand.get("name"):
                        df.at[df.index[i], "الماركة"] = brand["name"]
                    else:
                        # Try to extract brand from name (first word or two)
                        words = name.split()
                        guessed = words[0] if words else ""
                        if guessed and len(guessed) > 2:
                            df.at[df.index[i], "الماركة"] = guessed
                            existing = [b["اسم العلامة التجارية"] for b in st.session_state.new_brands]
                            if guessed not in existing and guessed not in new_brands_auto:
                                new_brands_auto.append(guessed)
                    if not str(df.iloc[i].get("تصنيف المنتج","")).strip():
                        df.at[df.index[i], "تصنيف المنتج"] = cat
                st.session_state.up_df = df
                if new_brands_auto:
                    for bn in new_brands_auto:
                        st.session_state.new_brands.append({
                            "اسم العلامة التجارية": bn,
                            "(SEO Page URL) رابط صفحة العلامة التجارية": to_slug(bn),
                            "وصف العلامة التجارية": "",
                            "صورة العلامة التجارية": "",
                        })
                    st.info(f"🆕 {len(new_brands_auto)} ماركة جديدة أُضيفت — صدّرها من الإعدادات")
                st.success(f"✅ تم التعيين لـ {len(idxs)} صف")
                st.rerun()

            st.divider()
            st.markdown("**تعديل يدوي لصف محدد:**")
            bc1, bc2, bc3 = st.columns(3)
            with bc1:
                b_row = st.number_input("رقم الصف", 0, max(0, len(df)-1), 0, key="b_row")
            bdf = st.session_state.brands_df
            cdf = st.session_state.categories_df
            brands_list = ["— اختر —"] + (
                [str(r.iloc[0]) for _, r in bdf.iterrows()] if bdf is not None else [])
            cats_list = ["— اختر —"] + (
                [str(r.get("التصنيفات","")) for _, r in cdf.iterrows()] if cdf is not None else [])
            with bc2:
                sel_brand = st.selectbox("الماركة", brands_list, key="sel_b")
            with bc3:
                sel_cat   = st.selectbox("التصنيف",  cats_list,   key="sel_c")
            if st.button("✅ تطبيق على الصف", key="apply_b"):
                if sel_brand != "— اختر —":
                    df.at[df.index[b_row], "الماركة"] = sel_brand
                if sel_cat != "— اختر —" and cdf is not None:
                    crow = cdf[cdf["التصنيفات"] == sel_cat]
                    if not crow.empty:
                        par  = str(crow.iloc[0].get("التصنيف الاساسي",""))
                        path = f"{par} > {sel_cat}" if par.strip() else sel_cat
                    else:
                        path = sel_cat
                    df.at[df.index[b_row], "تصنيف المنتج"] = path
                st.session_state.up_df = df
                st.success("✅ تم التطبيق")
                st.rerun()

        # ── Tab 3: Add Product ─────────────────────────────────
        with tabs[3]:
            st.markdown("**إضافة منتج جديد — أدخل الاسم وسيكمل النظام الباقي**")
            np1, np2, np3, np4 = st.columns(4)
            with np1: np_name   = st.text_input("اسم العطر ⭐", placeholder="ديور سوفاج 100 مل", key="np_nm")
            with np2: np_gender = st.selectbox("الجنس", ["للجنسين","للرجال","للنساء"], key="np_gn")
            with np3: np_size   = st.text_input("الحجم", "100 مل", key="np_sz")
            with np4: np_conc   = st.selectbox("التركيز", ["أو دو بارفيوم","أو دو كولون","أو دو تواليت","بارفيوم"], key="np_cn")
            np5, np6, np7, np8 = st.columns(4)
            with np5: np_price  = st.text_input("السعر", key="np_pr")
            with np6: np_sku    = st.text_input("SKU", key="np_sk")
            with np7: np_img    = st.text_input("رابط الصورة", key="np_im")
            with np8: np_type   = st.selectbox("النوع", ["عطر عادي","تستر"], key="np_tp")
            np9, np10 = st.columns(2)
            with np9:  np_weight = st.text_input("الوزن (kg)", "0.2", key="np_wt")
            with np10: np_brand_manual = st.text_input("الماركة (يدوي — اتركه فارغاً للكشف التلقائي)", key="np_br")

            ops1, ops2, ops3 = st.columns(3)
            with ops1: do_d = st.checkbox("🤖 توليد وصف AI",   value=True,  key="np_do_d")
            with ops2: do_i = st.checkbox("🖼 جلب صورة",       value=False, key="np_do_i")
            with ops3: do_s = st.checkbox("🔍 توليد SEO",       value=True,  key="np_do_s")

            if st.button("➕ إضافة للجدول", type="primary", key="add_to_table"):
                if not np_name.strip():
                    st.error("أدخل اسم العطر")
                else:
                    with st.spinner("جاري المعالجة..."):
                        is_t   = np_type == "تستر"
                        if np_brand_manual.strip():
                            brand = {"name": np_brand_manual.strip(),
                                     "page_url": to_slug(np_brand_manual.strip())}
                            # Check if new brand
                            existing = [b["اسم العلامة التجارية"] for b in st.session_state.new_brands]
                            if match_brand(np_name).get("name") == "" and np_brand_manual not in existing:
                                st.session_state.new_brands.append({
                                    "اسم العلامة التجارية": np_brand_manual.strip(),
                                    "(SEO Page URL) رابط صفحة العلامة التجارية": to_slug(np_brand_manual.strip()),
                                    "وصف العلامة التجارية": "",
                                    "صورة العلامة التجارية": "",
                                })
                        else:
                            brand  = match_brand(np_name)
                        cat    = match_category(np_name, np_gender)
                        seo    = gen_seo(np_name, brand, np_size, is_t, np_gender)
                        img    = np_img or (fetch_image(np_name, is_t) if do_i else "")
                        desc   = ai_generate(np_name, is_t, brand, np_size, np_gender, np_conc) \
                                 if do_d else ""
                        nr     = fill_row(name=np_name, price=np_price, sku=np_sku,
                                          image=img, desc=desc, brand=brand,
                                          category=cat, seo=seo, weight=np_weight)
                        new_df = pd.DataFrame([nr])
                        st.session_state.up_df = pd.concat(
                            [df, new_df], ignore_index=True)
                        if st.session_state.up_seo is not None:
                            st.session_state.up_seo = pd.concat([
                                st.session_state.up_seo,
                                pd.DataFrame([{
                                    "No. (غير قابل للتعديل)": "",
                                    "اسم المنتج (غير قابل للتعديل)": np_name,
                                    "رابط مخصص للمنتج (SEO Page URL)": seo["url"],
                                    "عنوان صفحة المنتج (SEO Page Title)": seo["title"],
                                    "وصف صفحة المنتج (SEO Page Description)": seo["desc"],
                                }])
                            ], ignore_index=True)
                    st.success(f"✅ تمت إضافة: **{np_name}**")
                    st.rerun()

        # ── Tab 4: Bulk Ops ────────────────────────────────────
        with tabs[4]:
            st.markdown("**تنفيذ عمليات متعددة دفعة واحدة على كل الصفوف**")
            bulk_ops = st.multiselect("اختر العمليات:", [
                "🏷 تعيين الماركات الفارغة",
                "📂 تعيين التصنيفات الفارغة",
                "📋 تعيين القيم الثابتة (نوع، شحن، ضريبة، وزن)",
                "🔤 توليد Alt Text للصور",
                "🔍 توليد SEO لكل الصفوف",
                "⚖️ تعيين وزن افتراضي للصفوف الفارغة",
            ], key="bulk_ops")

            dft_bulk_weight = st.text_input("الوزن الافتراضي (kg):", "0.2", key="bulk_wt")

            if st.button("⚡ تنفيذ الآن", type="primary", key="bulk_run"):
                prog = st.progress(0); stat = st.empty()
                seo_rows = []
                for n, (idx, row) in enumerate(df.iterrows()):
                    prog.progress(int((n+1)/len(df)*100))
                    name = str(row.get("أسم المنتج","")).strip()
                    if not name: continue
                    brand  = match_brand(name)
                    is_t   = any(w in name.lower() for w in ["تستر","tester"])
                    size_m = re.search(r"\d+\s*(?:مل|ml)", name, re.I)
                    size   = size_m.group() if size_m else "100 مل"
                    gender = ("للنساء" if any(w in name for w in ["نسائ","women"])
                              else "للرجال" if any(w in name for w in ["رجال","men"])
                              else "للجنسين")
                    seo = gen_seo(name, brand, size, is_t, gender)

                    if "🏷 تعيين الماركات الفارغة" in bulk_ops \
                            and not str(row.get("الماركة","")).strip():
                        df.at[idx, "الماركة"] = brand.get("name","")
                    if "📂 تعيين التصنيفات الفارغة" in bulk_ops \
                            and not str(row.get("تصنيف المنتج","")).strip():
                        df.at[idx, "تصنيف المنتج"] = match_category(name, gender)
                    if "📋 تعيين القيم الثابتة (نوع، شحن، ضريبة، وزن)" in bulk_ops:
                        df.at[idx, "النوع "]                    = "منتج"
                        df.at[idx, "نوع المنتج"]               = "منتج جاهز"
                        df.at[idx, "هل يتطلب شحن؟"]           = "نعم"
                        df.at[idx, "خاضع للضريبة ؟"]          = "نعم"
                        df.at[idx, "الوزن"]                    = df.at[idx, "الوزن"] or dft_bulk_weight
                        df.at[idx, "وحدة الوزن"]               = df.at[idx, "وحدة الوزن"] or "kg"
                        df.at[idx, "حالة المنتج"]              = df.at[idx, "حالة المنتج"] or "مرئي"
                        df.at[idx, "اقصي كمية لكل عميل"]      = df.at[idx, "اقصي كمية لكل عميل"] or "0"
                        df.at[idx, "إخفاء خيار تحديد الكمية"] = "0"
                        df.at[idx, "اضافة صورة عند الطلب"]    = "0"
                    if "⚖️ تعيين وزن افتراضي للصفوف الفارغة" in bulk_ops:
                        if not str(df.at[idx, "الوزن"]).strip() or str(df.at[idx, "الوزن"]).strip() in ("0","nan"):
                            df.at[idx, "الوزن"]      = dft_bulk_weight
                            df.at[idx, "وحدة الوزن"] = "kg"
                    if "🔤 توليد Alt Text للصور" in bulk_ops:
                        df.at[idx, "وصف صورة المنتج"] = seo["alt"]
                    if "🔍 توليد SEO لكل الصفوف" in bulk_ops:
                        seo_rows.append({
                            "No. (غير قابل للتعديل)":            str(row.get("No.","") or ""),
                            "اسم المنتج (غير قابل للتعديل)":     name,
                            "رابط مخصص للمنتج (SEO Page URL)":   seo["url"],
                            "عنوان صفحة المنتج (SEO Page Title)": seo["title"],
                            "وصف صفحة المنتج (SEO Page Description)": seo["desc"],
                        })

                st.session_state.up_df = df
                if seo_rows:
                    st.session_state.up_seo = pd.DataFrame(seo_rows)
                stat.markdown('<div class="prog-ok">✅ تمت جميع العمليات!</div>',
                              unsafe_allow_html=True)
                st.rerun()

        # ── EDITABLE GRID ─────────────────────────────────────
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>الجدول التفاعلي — عدّل أي خلية مباشرةً (مثل Excel)</h3></div>""",
                    unsafe_allow_html=True)

        all_c = list(df.columns)
        show_default = [c for c in EDITOR_COLS if c in all_c]
        show_cols = st.multiselect(
            "الأعمدة المعروضة:", options=all_c, default=show_default, key="show_cols")
        if not show_cols:
            show_cols = show_default or all_c[:8]

        edited = st.data_editor(
            df[show_cols].fillna(""),
            use_container_width=True,
            num_rows="dynamic",
            height=440,
            key="main_grid",
        )
        for c in show_cols:
            df[c] = edited[c]
        st.session_state.up_df = df

        # Description editor (single row)
        with st.expander("📝 تعديل الوصف HTML — منتج واحد"):
            sel_p = st.selectbox(
                "اختر المنتج:",
                range(len(df)),
                format_func=lambda i: str(df.iloc[i].get("أسم المنتج", f"صف {i}")),
                key="sel_p",
            )
            cur_d = str(df.iloc[sel_p].get("الوصف","") or "")
            new_d = st.text_area("الوصف (HTML):", value=cur_d, height=280, key="desc_area")
            if st.button("💾 حفظ الوصف", key="save_d"):
                df.at[df.index[sel_p], "الوصف"] = new_d
                st.session_state.up_df = df
                st.success("✅ تم حفظ الوصف")
                st.rerun()

        # SEO table
        if st.session_state.up_seo is not None:
            with st.expander("🔍 جدول SEO — قابل للتعديل"):
                ed_seo = st.data_editor(
                    st.session_state.up_seo.fillna(""),
                    use_container_width=True, num_rows="dynamic", key="seo_grid")
                st.session_state.up_seo = ed_seo

        # ── EXPORT ────────────────────────────────────────────
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>التصدير — جاهز للرفع على سلة</h3></div>""", unsafe_allow_html=True)

        e1, e2, e3, e4, e5 = st.columns(5)
        with e1:
            st.download_button(
                "📥 ملف المنتجات — Excel",
                export_product_xlsx(df),
                "mahwous_products.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="exp_px")
        with e2:
            st.download_button(
                "📥 ملف المنتجات — CSV",
                export_product_csv(df),
                "mahwous_products.csv", "text/csv",
                use_container_width=True, key="exp_pc")
        with e3:
            if st.session_state.up_seo is not None:
                st.download_button(
                    "📥 ملف SEO — Excel",
                    export_seo_xlsx(st.session_state.up_seo),
                    "mahwous_seo.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="exp_sx")
            else:
                st.info("نفّذ 'توليد SEO لكل الصفوف' أولاً")
        with e4:
            if st.session_state.up_seo is not None:
                st.download_button(
                    "📥 ملف SEO — CSV",
                    export_seo_csv(st.session_state.up_seo),
                    "mahwous_seo.csv", "text/csv",
                    use_container_width=True, key="exp_sc")
        with e5:
            if st.button("🔀 نقل للمقارنة", use_container_width=True, key="move_to_cmp"):
                st.session_state.cmp_new_df = df.copy()
                st.session_state.page = "compare"
                st.rerun()

    elif st.session_state.up_raw is None:
        st.markdown("""
        <div class="upload-zone">
          <div class="uz-icon">📂</div>
          <div class="uz-title">ارفع ملفك للبدء</div>
          <div class="uz-sub">يدعم: Excel (.xlsx / .xls) | CSV (UTF-8 / Windows-1256 / Latin)</div>
          <div class="uz-sub" style="margin-top:8px">
            ملفات سلة الجاهزة | ملفات الموردين | قوائم أسماء | أي تنسيق
          </div>
        </div>
        """, unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 1 — COMPARE v9.4 (مقارنة المنافسين)                      ║
# ╚══════════════════════════════════════════════════════════════════╝
if st.session_state.page == "compare_v2":

    st.markdown("""<div class="al-info">
    ارفع ملف متجرنا (مهووس) وملفات المنافسين (واحد أو أكثر) وملف الماركات الاختياري.
    سيقوم المحرك الذكي v9.4 بتحليل كل منتج منافس ومقارنته بمتجرنا باستخدام خوارزمية 5 طبقات
    (SKU + اسم نقي + حجم + تركيز + نوع) لاستخراج الفرص الجديدة الحقيقية.
    </div>""", unsafe_allow_html=True)

    # ── رفع الملفات ──────────────────────────────────────────────
    st.markdown("""<div class="sec-title"><div class="bar"></div><h3>رفع الملفات</h3></div>""",
                unsafe_allow_html=True)

    cv2_c1, cv2_c2, cv2_c3 = st.columns(3)

    with cv2_c1:
        st.markdown("**ملف متجرنا (مهووس)** — بكل الأعمدة")
        if st.session_state.cv2_store_df is not None:
            st.markdown(f'<div class="al-ok">محمّل: {len(st.session_state.cv2_store_df):,} منتج</div>',
                        unsafe_allow_html=True)
        up_cv2_store = st.file_uploader("ارفع ملف المتجر", type=["csv","xlsx","xls"],
                                         key="cv2_store_up", label_visibility="collapsed")
        if up_cv2_store:
            df_s = read_file(up_cv2_store, salla_2row=True)
            if df_s.empty:
                df_s = read_file(up_cv2_store, salla_2row=False)
            if not df_s.empty:
                st.session_state.cv2_store_df = df_s
                st.success(f"✅ {len(df_s):,} منتج في المتجر")

    with cv2_c2:
        st.markdown("**ملفات المنافسين** — يمكن رفع أكثر من ملف")
        if st.session_state.cv2_comp_dfs:
            total_comp = sum(len(d) for d in st.session_state.cv2_comp_dfs)
            st.markdown(f'<div class="al-ok">محمّل: {total_comp:,} منتج من {len(st.session_state.cv2_comp_dfs)} ملف</div>',
                        unsafe_allow_html=True)
        up_cv2_comp = st.file_uploader("ارفع ملفات المنافسين", type=["csv","xlsx","xls"],
                                        key="cv2_comp_up", accept_multiple_files=True,
                                        label_visibility="collapsed")
        if up_cv2_comp:
            new_dfs = []
            for f in up_cv2_comp:
                df_c = read_file(f)
                if not df_c.empty:
                    df_c["_source"] = f.name
                    new_dfs.append(df_c)
            if new_dfs:
                st.session_state.cv2_comp_dfs = new_dfs
                total = sum(len(d) for d in new_dfs)
                st.success(f"✅ {total:,} منتج من {len(new_dfs)} ملف")

    with cv2_c3:
        st.markdown("**ملف الماركات** (اختياري)")
        if st.session_state.cv2_brands_df is not None:
            st.markdown(f'<div class="al-ok">محمّل: {len(st.session_state.cv2_brands_df):,} ماركة</div>',
                        unsafe_allow_html=True)
        elif st.session_state.brands_df is not None:
            st.markdown(f'<div class="al-ok">يستخدم ملف الماركات الافتراضي: {len(st.session_state.brands_df):,} ماركة</div>',
                        unsafe_allow_html=True)
        up_cv2_brands = st.file_uploader("ارفع ملف الماركات", type=["csv","xlsx"],
                                          key="cv2_brands_up", label_visibility="collapsed")
        if up_cv2_brands:
            df_br = read_file(up_cv2_brands)
            if not df_br.empty:
                st.session_state.cv2_brands_df = df_br
                st.success(f"✅ {len(df_br):,} ماركة")

    # ── مؤشر الحالة ──────────────────────────────────────────────
    has_store = st.session_state.cv2_store_df is not None
    has_comp  = bool(st.session_state.cv2_comp_dfs)
    if has_store and not has_comp:
        st.info("✅ ملف المتجر جاهز — الآن ارفع ملفات المنافسين للبدء")
    elif not has_store and has_comp:
        st.info("✅ ملفات المنافسين جاهزة — الآن ارفع ملف متجرنا للبدء")
    elif not has_store and not has_comp:
        st.markdown("""<div class="upload-zone"><div class="uz-icon">🔎</div>
        <div class="uz-title">ارفع ملف متجرنا وملفات المنافسين للبدء</div>
        <div class="uz-sub">المحرك الذكي v9.4 يستخدم 5 طبقات مقارنة: SKU + اسم نقي + حجم + تركيز + نوع</div>
        </div>""", unsafe_allow_html=True)

    # ── إعدادات المحرك ───────────────────────────────────────────
    if has_store and has_comp:
        store_df_v2 = st.session_state.cv2_store_df
        comp_dfs_v2 = st.session_state.cv2_comp_dfs

        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>إعدادات المحرك</h3></div>""", unsafe_allow_html=True)

        # تعيين الأعمدة تلقائياً
        NONE_V2 = "— لا يوجد —"
        store_opts_v2 = [NONE_V2] + list(store_df_v2.columns)
        comp_all_cols = list(pd.concat(comp_dfs_v2, ignore_index=True).columns)
        comp_opts_v2  = [NONE_V2] + comp_all_cols

        def giv2(cols, kws, opts):
            g = auto_guess_col(cols, kws)
            return opts.index(g) if g in opts else 0

        cv2_r1, cv2_r2, cv2_r3, cv2_r4 = st.columns(4)
        with cv2_r1:
            store_nm_col = st.selectbox("عمود الاسم (المتجر):", store_opts_v2,
                index=giv2(store_df_v2.columns, ["اسم","name","منتج"], store_opts_v2),
                key="cv2_snm")
        with cv2_r2:
            store_sk_col = st.selectbox("عمود SKU (المتجر):", store_opts_v2,
                index=giv2(store_df_v2.columns, ["sku","رمز","barcode"], store_opts_v2),
                key="cv2_ssk")
        with cv2_r3:
            comp_nm_col = st.selectbox("عمود الاسم (المنافس):", comp_opts_v2,
                index=giv2(comp_all_cols, ["اسم","name","منتج"], comp_opts_v2),
                key="cv2_cnm")
        with cv2_r4:
            comp_pr_col = st.selectbox("عمود السعر (المنافس):", comp_opts_v2,
                index=giv2(comp_all_cols, ["سعر","price","text-sm"], comp_opts_v2),
                key="cv2_cpr")

        cv2_r5, cv2_r6, cv2_r7 = st.columns(3)
        with cv2_r5:
            comp_img_col = st.selectbox("عمود الصورة (المنافس):", comp_opts_v2,
                index=giv2(comp_all_cols, ["صورة","src","image"], comp_opts_v2),
                key="cv2_cimg")
        with cv2_r6:
            t_dup_v2  = st.slider("عتبة المكرر (%):", 70, 98, 88, key="cv2_tdup")
        with cv2_r7:
            t_near_v2 = st.slider("عتبة المراجعة (%):", 40, 85, 60, key="cv2_tnear")

        if not HAS_RAPIDFUZZ:
            st.warning("⚠️ rapidfuzz غير مثبّت — يعمل بخوارزمية بديلة أقل دقة. أضف `rapidfuzz` إلى requirements.txt للحصول على أعلى دقة.")

        if st.button("🚀 تشغيل المحرك الذكي v9.4", type="primary", key="run_cv2"):
            if store_nm_col == NONE_V2:
                st.error("حدد عمود اسم المنتج في ملف المتجر")
            elif comp_nm_col == NONE_V2:
                st.error("حدد عمود اسم المنتج في ملفات المنافسين")
            else:
                # دمج ملفات المنافسين
                comp_merged = pd.concat(comp_dfs_v2, ignore_index=True)

                # استخراج قائمة الماركات
                brands_v2 = []
                bdf_v2 = st.session_state.cv2_brands_df or st.session_state.brands_df
                if bdf_v2 is not None:
                    col0 = bdf_v2.columns[0]
                    brands_v2 = bdf_v2[col0].dropna().astype(str).str.strip().tolist()

                with st.spinner(f"جاري تحليل {len(comp_merged):,} منتج من المنافسين..."):
                    results_v2 = run_smart_comparison(
                        new_df=comp_merged,
                        store_df=store_df_v2,
                        new_name_col=comp_nm_col,
                        store_name_col=store_nm_col,
                        new_sku_col=None,
                        store_sku_col=store_sk_col if store_sk_col != NONE_V2 else None,
                        new_img_col=comp_img_col if comp_img_col != NONE_V2 else None,
                        t_dup=t_dup_v2,
                        t_near=t_near_v2,
                        t_review=40,
                        brands_list=brands_v2,
                    )
                    # إضافة عمود السعر من ملف المنافس
                    if comp_pr_col != NONE_V2 and comp_pr_col in comp_merged.columns:
                        price_map = {i: str(comp_merged.iloc[i].get(comp_pr_col, ""))
                                     for i in range(len(comp_merged))}
                        results_v2["سعر المنافس"] = results_v2["_idx"].map(
                            lambda x: price_map.get(x, ""))
                    st.session_state.cv2_results = results_v2
                st.rerun()

    # ── عرض النتائج ──────────────────────────────────────────────
    if st.session_state.cv2_results is not None:
        res_v2 = st.session_state.cv2_results

        new_opps  = res_v2[res_v2["الحالة"] == "جديد"]
        dups_v2   = res_v2[res_v2["الحالة"] == "مكرر"]
        reviews_v2 = res_v2[res_v2["الحالة"] == "مشبوه"]

        st.markdown(f"""
        <div class="stats-bar">
          <div class="stat-box"><div class="n">{len(res_v2):,}</div><div class="lb">إجمالي المنتجات</div></div>
          <div class="stat-box"><div class="n" style="color:#43a047">{len(new_opps):,}</div><div class="lb">فرص جديدة</div></div>
          <div class="stat-box"><div class="n" style="color:#e53935">{len(dups_v2):,}</div><div class="lb">مكررات</div></div>
          <div class="stat-box"><div class="n" style="color:#f9a825">{len(reviews_v2):,}</div><div class="lb">تحتاج مراجعة</div></div>
        </div>
        """, unsafe_allow_html=True)

        # معاينة أول 20 فرصة جديدة
        if not new_opps.empty:
            st.markdown("""<div class="sec-title"><div class="bar"></div>
            <h3>معاينة الفرص الجديدة (أول 20)</h3></div>""", unsafe_allow_html=True)
            preview_cols = ["الاسم الجديد", "الماركة", "التصنيف", "نسبة التشابه", "سبب القرار"]
            if "سعر المنافس" in new_opps.columns:
                preview_cols.insert(2, "سعر المنافس")
            st.dataframe(new_opps[preview_cols].head(20), use_container_width=True)

        # أزرار التصدير
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>تصدير النتائج</h3></div>""", unsafe_allow_html=True)

        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")

        def _build_salla_from_results(df_results, comp_merged_df, comp_nm_col, comp_img_col, comp_pr_col):
            """بناء ملف سلة من نتائج المقارنة."""
            rows = []
            for _, r in df_results.iterrows():
                idx = r["_idx"]
                if idx < len(comp_merged_df):
                    orig_row = comp_merged_df.iloc[idx]
                    img  = str(orig_row.get(comp_img_col, "") or "") if comp_img_col and comp_img_col != NONE_V2 else r.get("_img", "")
                    price = str(orig_row.get(comp_pr_col, "") or "") if comp_pr_col and comp_pr_col != NONE_V2 else ""
                else:
                    img = r.get("_img", "")
                    price = ""
                row = {col: "" for col in SALLA_COLS}
                row["النوع "]         = "منتج"
                row["أسم المنتج"]     = r["الاسم الجديد"]
                row["تصنيف المنتج"]   = r.get("التصنيف", "العطور")
                row["صورة المنتج"]    = img
                row["وصف صورة المنتج"] = r["الاسم الجديد"]
                row["نوع المنتج"]     = "منتج جاهز"
                row["سعر المنتج"]     = price
                row["هل يتطلب شحن؟"] = "نعم"
                row["الوزن"]          = "0.2"
                row["وحدة الوزن"]     = "kg"
                row["الماركة"]        = r.get("الماركة", "")
                row["خاضع للضريبة ؟"] = "نعم"
                row["اقصي كمية لكل عميل"] = "0"
                row["تثبيت المنتج"]   = "لا"
                row["اضافة صورة عند الطلب"] = "لا"
                rows.append(row)
            return pd.DataFrame(rows, columns=SALLA_COLS)

        comp_merged_export = pd.concat(st.session_state.cv2_comp_dfs, ignore_index=True)                              if st.session_state.cv2_comp_dfs else pd.DataFrame()

        exp1, exp2, exp3 = st.columns(3)
        with exp1:
            if not new_opps.empty:
                salla_new = _build_salla_from_results(
                    new_opps, comp_merged_export,
                    comp_nm_col if "comp_nm_col" in dir() else "الاسم الجديد",
                    comp_img_col if "comp_img_col" in dir() else None,
                    comp_pr_col if "comp_pr_col" in dir() else None,
                )
                st.download_button(
                    f"📥 الفرص الجديدة — سلة ({len(new_opps):,} منتج)",
                    export_product_csv(salla_new),
                    f"منتج_جديد_{date_str}.csv", "text/csv",
                    use_container_width=True, key="dl_cv2_new"
                )
        with exp2:
            if not res_v2.empty:
                report_cols = ["الاسم الجديد", "الماركة", "التصنيف",
                               "أقرب تطابق في المتجر", "نسبة التشابه",
                               "الحالة", "سبب القرار"]
                if "سعر المنافس" in res_v2.columns:
                    report_cols.insert(3, "سعر المنافس")
                report_df = res_v2[[c for c in report_cols if c in res_v2.columns]]
                st.download_button(
                    "📊 تقرير كامل — CSV",
                    report_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig"),
                    f"تقرير_مقارنة_{date_str}.csv", "text/csv",
                    use_container_width=True, key="dl_cv2_rep"
                )
        with exp3:
            if not new_opps.empty:
                if st.button("🛠️ نقل الفرص الجديدة للمُعالج", key="cv2_to_proc", use_container_width=True):
                    salla_new2 = _build_salla_from_results(
                        new_opps, comp_merged_export,
                        comp_nm_col if "comp_nm_col" in dir() else "الاسم الجديد",
                        comp_img_col if "comp_img_col" in dir() else None,
                        comp_pr_col if "comp_pr_col" in dir() else None,
                    )
                    st.session_state.up_df      = salla_new2
                    st.session_state.up_mapped  = True
                    st.session_state.up_filename = f"فرص_جديدة_{date_str}.csv"
                    st.session_state.page       = "processor"
                    st.rerun()

        if st.button("🔄 إعادة ضبط المحرك", key="reset_cv2"):
            st.session_state.cv2_results  = None
            st.session_state.cv2_store_df = None
            st.session_state.cv2_comp_dfs = []
            st.rerun()

    # (empty zone now shown above via has_store/has_comp logic)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 3 — QUICK ADD                                             ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "quickadd":

    st.markdown("""<div class="al-info">
    أضف منتجات جديدة بسرعة بطريقتين: (1) من خلال رابط منتج أو أكثر، أو (2) بإدخال يدوي سريع مع رفع صور.
    </div>""", unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["🔗 سحب من رابط", "📝 إدخال يدوي ورفع صور"])

    # ── TAB 1: Fetch from URLs (multiple) ───────────────────────
    with tab1:
        st.markdown("### سحب بيانات المنتجات من روابط")
        st.caption("أدخل رابطاً واحداً أو أكثر — كل رابط في سطر منفصل")

        # Initialize URL list in session state
        if "qa_url_list" not in st.session_state:
            st.session_state.qa_url_list = [""]

        # Dynamic URL inputs
        urls_to_remove = []
        for idx, url_val in enumerate(st.session_state.qa_url_list):
            col_url, col_del = st.columns([10, 1])
            with col_url:
                new_val = st.text_input(
                    f"رابط المنتج {idx + 1}",
                    value=url_val,
                    placeholder="https://example.com/product/...",
                    key=f"qa_url_{idx}",
                    label_visibility="collapsed",
                )
                st.session_state.qa_url_list[idx] = new_val
            with col_del:
                if len(st.session_state.qa_url_list) > 1:
                    if st.button("✕", key=f"del_url_{idx}", help="حذف هذا الرابط"):
                        urls_to_remove.append(idx)

        for idx in reversed(urls_to_remove):
            st.session_state.qa_url_list.pop(idx)
            st.rerun()

        col_add, col_fetch = st.columns([1, 3])
        with col_add:
            if st.button("➕ إضافة رابط آخر", use_container_width=True):
                st.session_state.qa_url_list.append("")
                st.rerun()
        with col_fetch:
            do_fetch = st.button("🔄 سحب البيانات من الروابط", type="primary",
                                 use_container_width=True, key="qa_fetch_urls")

        # Options for URL scraping
        uo1, uo2 = st.columns(2)
        with uo1:
            qa_url_gen_desc = st.checkbox("🤖 توليد وصف AI لكل منتج", value=True, key="qa_url_gen_desc")
        with uo2:
            qa_url_gen_seo = st.checkbox("🔍 توليد SEO", value=True, key="qa_url_gen_seo")

        if do_fetch:
            valid_urls = [u.strip() for u in st.session_state.qa_url_list if u.strip()]
            if not valid_urls:
                st.error("الرجاء إدخال رابط منتج واحد على الأقل")
            else:
                progress_bar = st.progress(0, text="جاري السحب...")
                success_count = 0
                for url_i, url_item in enumerate(valid_urls):
                    progress_bar.progress(
                        int((url_i / len(valid_urls)) * 100),
                        text=f"جاري سحب المنتج {url_i + 1} من {len(valid_urls)}: {url_item[:60]}..."
                    )
                    with st.spinner(f"سحب: {url_item[:80]}..."):
                        scraped = scrape_product_url(url_item)

                    if scraped.get("error"):
                        st.warning(f"⚠️ تعذّر سحب {url_item[:60]}: {scraped['error']}")
                        continue

                    ex_name  = scraped.get("name", "") or "منتج جديد"
                    ex_price = scraped.get("price", "") or ""
                    ex_img   = scraped.get("image", "") or ""
                    ex_imgs  = scraped.get("images", [])
                    ex_desc  = scraped.get("desc", "") or ""
                    ex_brand_hint = scraped.get("brand_hint", "") or ""

                    # Show preview card
                    with st.expander(f"📦 {ex_name[:80]}", expanded=True):
                        pc1, pc2 = st.columns([1, 3])
                        with pc1:
                            if ex_img:
                                st.image(ex_img, width=120, caption="الصورة الرئيسية")
                            if len(ex_imgs) > 1:
                                st.caption(f"📷 {len(ex_imgs)} صور متاحة")
                                thumb_cols = st.columns(min(len(ex_imgs), 4))
                                for ti, timg in enumerate(ex_imgs[:4]):
                                    with thumb_cols[ti]:
                                        try:
                                            st.image(timg, width=60)
                                        except Exception:
                                            pass
                        with pc2:
                            st.markdown(f"**الاسم:** {ex_name}")
                            st.markdown(f"**السعر:** {ex_price} ريال" if ex_price else "**السعر:** غير محدد")
                            st.markdown(f"**الماركة المكتشفة:** {ex_brand_hint}" if ex_brand_hint else "")
                            st.caption(f"**الوصف:** {ex_desc[:200]}..." if len(ex_desc) > 200 else f"**الوصف:** {ex_desc}")

                    # Match brand
                    if ex_brand_hint:
                        brand = match_brand(ex_brand_hint)
                        if not brand.get("name"):
                            brand = generate_new_brand(ex_brand_hint)
                            existing_b = [b.get("اسم العلامة التجارية", "") for b in st.session_state.new_brands]
                            if ex_brand_hint not in existing_b:
                                st.session_state.new_brands.append({
                                    "اسم العلامة التجارية": brand.get("name", ex_brand_hint),
                                    "(SEO Page URL) رابط صفحة العلامة التجارية": brand.get("page_url", to_slug(ex_brand_hint)),
                                    "وصف العلامة التجارية": brand.get("desc", ""),
                                    "صورة العلامة التجارية": "",
                                })
                    else:
                        brand = match_brand(ex_name)

                    cat = match_category(ex_name, "للجنسين")

                    # Extract size from name
                    size_match = re.search(r'(\d+)\s*(?:ml|مل|ML)', ex_name, re.IGNORECASE)
                    ex_size = size_match.group(0) if size_match else "100 مل"

                    seo = gen_seo(ex_name, brand, ex_size, False, "للجنسين") if qa_url_gen_seo else {"url": "", "title": "", "desc": ""}

                    # Generate AI description
                    if qa_url_gen_desc and st.session_state.api_key:
                        final_desc = ai_generate(ex_name, False, brand, ex_size, "للجنسين", "أو دو بارفيوم")
                    else:
                        final_desc = f"<p>{ex_desc}</p>" if ex_desc else f"<p>وصف مبدئي لـ {ex_name}</p>"

                    nr = fill_row(
                        name=ex_name, price=ex_price, image=ex_img,
                        desc=final_desc, brand=brand, category=cat,
                        seo=seo, weight="0.2"
                    )
                    st.session_state.qa_rows.append({
                        "product": nr,
                        "seo": {"url": seo["url"], "title": seo["title"], "desc": seo["desc"]},
                        "images": ex_imgs,
                    })
                    success_count += 1

                progress_bar.progress(100, text="اكتمل السحب!")
                if success_count:
                    st.success(f"✅ تم سحب {success_count} منتج بنجاح!")
                    # Reset URL list
                    st.session_state.qa_url_list = [""]
                    st.rerun()

    # ── TAB 2: Manual Entry with Image Upload ────────────────────
    with tab2:
        with st.form("qa_form", clear_on_submit=True):
            f1, f2, f3 = st.columns(3)
            with f1:
                qa_nm = st.text_input("اسم العطر ⭐", placeholder="مثال: شانيل بلو دو شانيل 100 مل للرجال")
                qa_pr = st.text_input("السعر ⭐", placeholder="299")
            with f2:
                qa_gn = st.selectbox("الجنس", ["للجنسين","للرجال","للنساء"])
                qa_sk = st.text_input("SKU", placeholder="اختياري")
            with f3:
                qa_sz = st.text_input("الحجم", "100 مل")
                qa_cn = st.selectbox("التركيز", ["أو دو بارفيوم","أو دو كولون","أو دو تواليت","بارفيوم"])

            f4, f5, f6 = st.columns(3)
            with f4: qa_tp   = st.selectbox("النوع", ["عطر عادي","تستر"])
            with f5: qa_wt   = st.text_input("الوزن (kg)", "0.2")
            with f6: qa_br   = st.text_input("الماركة (اختياري)")
            
            st.markdown("**صور المنتج**")
            qa_imgs = st.file_uploader("ارفع صورة أو صورتين", type=["png","jpg","jpeg","webp"], accept_multiple_files=True)

            o1, o2, o3 = st.columns(3)
            with o1: qa_do_d = st.checkbox("🤖 توليد وصف AI",   value=True)
            with o2: qa_do_i = st.checkbox("🖼 جلب صورة من جوجل (إذا لم ترفع)", value=False)
            with o3: qa_do_s = st.checkbox("🔍 توليد SEO",       value=True)

            sub = st.form_submit_button("➕ إضافة للقائمة وتجهيز الملف", type="primary", use_container_width=True)

        if sub:
            if not qa_nm.strip() or not qa_pr.strip():
                st.error("الاسم والسعر حقول إجبارية!")
            else:
                with st.spinner("جاري التجهيز..."):
                    is_t   = qa_tp == "تستر"
                    if qa_br.strip():
                        brand = {"name": qa_br.strip(), "page_url": to_slug(qa_br.strip())}
                        existing_brands = [b["اسم العلامة التجارية"] for b in st.session_state.new_brands]
                        if match_brand(qa_nm).get("name") == "" and qa_br not in existing_brands:
                            st.session_state.new_brands.append({
                                "اسم العلامة التجارية": qa_br.strip(),
                                "(SEO Page URL) رابط صفحة العلامة التجارية": to_slug(qa_br.strip()),
                                "وصف العلامة التجارية": "",
                                "صورة العلامة التجارية": "",
                            })
                    else:
                        brand  = match_brand(qa_nm)
                    
                    cat    = match_category(qa_nm, qa_gn)
                    seo    = gen_seo(qa_nm, brand, qa_sz, is_t, qa_gn)
                    
                    # Handle Images
                    img_url = ""
                    if qa_imgs:
                        img_url = f"https://mahwous.com/uploads/{qa_imgs[0].name}" # Mock URL for uploaded image
                    elif qa_do_i:
                        img_url = fetch_image(qa_nm, is_t)
                        
                    desc   = ai_generate(qa_nm, is_t, brand, qa_sz, qa_gn, qa_cn) if qa_do_d else ""
                    nr     = fill_row(name=qa_nm, price=qa_pr, sku=qa_sk, image=img_url,
                                      desc=desc, brand=brand, category=cat, seo=seo,
                                      weight=qa_wt)
                    st.session_state.qa_rows.append({
                        "product": nr,
                        "seo": {"url": seo["url"], "title": seo["title"], "desc": seo["desc"]},
                    })
                st.success(f"✅ تمت الإضافة: **{qa_nm}**")

    # ── Common List Display ───────────────────────────────────────
    if st.session_state.qa_rows:
        st.markdown(f"### القائمة ({len(st.session_state.qa_rows)} منتج)")
        prev = []
        for r in st.session_state.qa_rows:
            p = r["product"]
            prev.append({
                "الاسم":    p.get("أسم المنتج",""),
                "الماركة":  p.get("الماركة",""),
                "التصنيف":  p.get("تصنيف المنتج",""),
                "السعر":    p.get("سعر المنتج",""),
                "الوزن":    p.get("الوزن",""),
                "وصف ✓":   "✅" if str(p.get("الوصف","")).strip() else "—",
                "صورة ✓":  "✅" if str(p.get("صورة المنتج","")).startswith("http") else "—",
            })
        st.dataframe(pd.DataFrame(prev), use_container_width=True)

        prod_df_qa = pd.DataFrame([r["product"] for r in st.session_state.qa_rows])
        seo_df_qa  = pd.DataFrame([{
            "No. (غير قابل للتعديل)": "",
            "اسم المنتج (غير قابل للتعديل)": r["product"]["أسم المنتج"],
            "رابط مخصص للمنتج (SEO Page URL)": r["seo"]["url"],
            "عنوان صفحة المنتج (SEO Page Title)": r["seo"]["title"],
            "وصف صفحة المنتج (SEO Page Description)": r["seo"]["desc"],
        } for r in st.session_state.qa_rows])

        qe1, qe2, qe3, qe4, qe5 = st.columns(5)
        with qe1:
            st.download_button("📥 منتجات Excel",
                export_product_xlsx(prod_df_qa), "qa_products.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with qe2:
            st.download_button("📥 منتجات CSV",
                export_product_csv(prod_df_qa), "qa_products.csv", "text/csv",
                use_container_width=True)
        with qe3:
            st.download_button("📥 SEO Excel",
                export_seo_xlsx(seo_df_qa), "qa_seo.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with qe4:
            st.download_button("📥 SEO CSV",
                export_seo_csv(seo_df_qa), "qa_seo.csv", "text/csv",
                use_container_width=True)
        with qe5:
            if st.button("🔀 نقل لتجهيز الملفات", use_container_width=True, key="move_qa"):
                ex = st.session_state.up_df
                st.session_state.up_df = pd.concat(
                    [ex, prod_df_qa], ignore_index=True) if ex is not None else prod_df_qa
                st.session_state.up_mapped   = True
                st.session_state.up_filename = "منتجات سريعة"
                st.session_state.qa_rows     = []
                st.session_state.page        = "processor"
                st.rerun()

        if st.button("🗑️ مسح القائمة", key="clear_qa"):
            st.session_state.qa_rows = []
            st.rerun()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 4 — COMPARE & DEDUP (NEW)                                 ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "compare":

    st.markdown("""<div class="al-info">
    ارفع ملف المنتجات الجديدة (المُعالج) وملف المتجر الأساسي. سيقارن النظام المنتجات
    ويستبعد المكرر تلقائياً، ويعرض المنتجات المشبوهة (تشابه 60-99%) لتقرر اعتمادها أو إلغاءها.
    </div>""", unsafe_allow_html=True)

    # ── Upload Section ────────────────────────────────────────────
    st.markdown("""<div class="sec-title"><div class="bar"></div><h3>رفع الملفات</h3></div>""",
                unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**ملف المنتجات الجديدة** (المُعالج من المُعالج الشامل)")
        if st.session_state.cmp_new_df is not None:
            st.markdown(f'<div class="al-ok">محمّل من المُعالج: {len(st.session_state.cmp_new_df)} منتج</div>',
                        unsafe_allow_html=True)
        up_new = st.file_uploader("أو ارفع ملف جديد", type=["csv","xlsx","xls"],
                                   key="cmp_new_up", label_visibility="collapsed")
        if up_new:
            df_new = read_file(up_new)
            if not df_new.empty:
                st.session_state.cmp_new_df = df_new
                st.success(f"✅ {len(df_new)} منتج جديد")

    with col_b:
        st.markdown("**ملف المتجر الأساسي** (متجرنا مهووس بكل الأعمدة)")
        if st.session_state.cmp_store_df is not None:
            st.markdown(f'<div class="al-ok">محمّل: {len(st.session_state.cmp_store_df)} منتج في المتجر</div>',
                        unsafe_allow_html=True)
        up_store = st.file_uploader("ارفع ملف المتجر الأساسي", type=["csv","xlsx","xls"],
                                     key="cmp_store_up", label_visibility="collapsed")
        if up_store:
            df_store = read_file(up_store, salla_2row=True)
            if not df_store.empty:
                st.session_state.cmp_store_df = df_store
                st.success(f"✅ {len(df_store)} منتج في المتجر")

    # ── Column Mapping for Comparison ────────────────────────────
    if st.session_state.cmp_new_df is not None and st.session_state.cmp_store_df is not None:
        new_df   = st.session_state.cmp_new_df
        store_df = st.session_state.cmp_store_df

        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>تعيين أعمدة المقارنة</h3></div>""", unsafe_allow_html=True)

        NONE_C = "— لا يوجد —"
        new_opts   = [NONE_C] + list(new_df.columns)
        store_opts = [NONE_C] + list(store_df.columns)

        def gi2(cols, kws):
            g = auto_guess_col(cols, kws)
            opts = [NONE_C] + list(cols)
            return opts.index(g) if g in opts else 0

        cm1, cm2, cm3, cm4 = st.columns(4)
        with cm1:
            new_name_col = st.selectbox("عمود الاسم (الجديد):", new_opts,
                index=gi2(new_df.columns, ["اسم","name","منتج","عطر"]), key="cmp_nm_new")
        with cm2:
            new_sku_col  = st.selectbox("عمود SKU (الجديد):", new_opts,
                index=gi2(new_df.columns, ["sku","رمز","barcode"]), key="cmp_sk_new")
        with cm3:
            store_name_col = st.selectbox("عمود الاسم (المتجر):", store_opts,
                index=gi2(store_df.columns, ["اسم","name","منتج","عطر"]), key="cmp_nm_st")
        with cm4:
            store_sku_col  = st.selectbox("عمود SKU (المتجر):", store_opts,
                index=gi2(store_df.columns, ["sku","رمز","barcode"]), key="cmp_sk_st")

        sim_threshold = st.slider("حد التشابه للمنتجات المشبوهة (%):", 50, 95, 75, key="sim_thr")

        if st.button("🔍 تشغيل المقارنة الآن", type="primary", key="run_cmp"):
            if new_name_col == NONE_C:
                st.error("حدد عمود اسم المنتج في الملف الجديد")
            elif store_name_col == NONE_C:
                st.error("حدد عمود اسم المنتج في ملف المتجر")
            else:
                with st.spinner("جاري المقارنة..."):
                    # Build store name & SKU sets
                    store_names = [str(v).strip().lower() for v in
                                   store_df[store_name_col].fillna("").tolist() if str(v).strip()]
                    store_skus  = set()
                    if store_sku_col != NONE_C:
                        store_skus = {str(v).strip().lower() for v in
                                      store_df[store_sku_col].fillna("").tolist() if str(v).strip()}

                    # ── Advanced 5-Stage Comparison Engine ──
                    results = []
                    
                    # Pre-process store products for advanced matching
                    store_parsed = []
                    for idx, row in store_df.iterrows():
                        sname = str(row.get(store_name_col, "") or "").strip()
                        if not sname: continue
                        attrs = extract_product_attrs(sname)
                        store_parsed.append({
                            "orig_name": sname,
                            "clean_name": attrs["clean_name"],
                            "size": attrs["size"],
                            "type": attrs["type"],
                            "concentration": attrs["concentration"],
                            "sku": str(row.get(store_sku_col, "") or "").strip() if store_sku_col != NONE_C else ""
                        })
                    
                    store_clean_dict = {i: p["clean_name"] for i, p in enumerate(store_parsed)}
                    store_skus = {p["sku"].lower(): p["orig_name"] for p in store_parsed if p["sku"]}

                    for i, row in new_df.iterrows():
                        new_name = str(row.get(new_name_col, "") or "").strip()
                        new_sku  = str(row.get(new_sku_col, "") or "").strip() if new_sku_col != NONE_C else ""
                        new_img  = str(row.get("صورة المنتج","") or "")
                        
                        if not new_name:
                            continue

                        # Stage 1: Exact SKU Match
                        if new_sku and new_sku.lower() in store_skus:
                            results.append({
                                "الاسم الجديد":      new_name,
                                "SKU الجديد":        new_sku,
                                "أقرب تطابق في المتجر": store_skus[new_sku.lower()],
                                "نسبة التشابه":      100,
                                "الحالة":            "مكرر (SKU)",
                                "الإجراء":           "حذف",
                                "_idx":              i,
                                "_img":              new_img,
                            })
                            continue

                        # Stage 2 & 3: Advanced Attribute & Fuzzy Matching
                        new_attrs = extract_product_attrs(new_name)
                        new_clean = new_attrs["clean_name"]
                        new_size  = new_attrs["size"]
                        new_type  = new_attrs["type"]
                        new_conc  = new_attrs["concentration"]
                        
                        best_score = 0
                        best_match = ""
                        best_sp = None
                        
                        if store_clean_dict:
                            if HAS_RAPIDFUZZ:
                                best = rf_process.extractOne(new_clean, store_clean_dict, scorer=rf_fuzz.token_set_ratio)
                                if best:
                                    _, best_score, pos = best
                                    best_sp = store_parsed[pos]
                                    best_match = best_sp["orig_name"]
                            else:
                                for idx2, clean2 in store_clean_dict.items():
                                    s = _fuzzy_ratio(new_clean, clean2)
                                    if s > best_score:
                                        best_score = s
                                        best_sp = store_parsed[idx2]
                                        best_match = best_sp["orig_name"]

                        # Stage 4: Logical Verdict based on attributes
                        if best_score >= 90:
                            # High text similarity - check attributes
                            if new_type != best_sp["type"]:
                                status = "جديد"
                                action = "اعتماد"
                                reason = f"نوع مختلف: {new_type} vs {best_sp['type']}"
                                best_score -= 20 # Penalize score for display
                            elif new_size != best_sp["size"] and new_size != 0 and best_sp["size"] != 0:
                                status = "جديد"
                                action = "اعتماد"
                                reason = f"حجم مختلف: {new_size}ml vs {best_sp['size']}ml"
                                best_score -= 15
                            elif new_conc != best_sp["concentration"] and new_conc != "غير محدد" and best_sp["concentration"] != "غير محدد":
                                status = "جديد"
                                action = "اعتماد"
                                reason = f"تركيز مختلف: {new_conc} vs {best_sp['concentration']}"
                                best_score -= 10
                            else:
                                status = "مكرر (اسم وخصائص)"
                                action = "حذف"
                        elif best_score >= sim_threshold:
                            status = "مشبوه"
                            action = "مراجعة"
                        else:
                            status = "جديد"
                            action = "اعتماد"

                        results.append({
                            "الاسم الجديد":      new_name,
                            "SKU الجديد":        new_sku,
                            "أقرب تطابق في المتجر": best_match,
                            "نسبة التشابه":      best_score,
                            "الحالة":            status,
                            "الإجراء":           action,
                            "_idx":              i,
                            "_img":              new_img,
                        })

                    st.session_state.cmp_results  = pd.DataFrame(results)
                    st.session_state.cmp_approved = {
                        r["_idx"]: (r["الإجراء"] == "اعتماد")
                        for r in results
                    }
                st.rerun()

    # ── Show Results ──────────────────────────────────────────────
    if st.session_state.cmp_results is not None:
        res = st.session_state.cmp_results

        exact_dup  = res[res["الحالة"].str.contains("مكرر")]
        suspect    = res[res["الحالة"] == "مشبوه"]
        new_clean  = res[res["الحالة"] == "جديد"]

        st.markdown(f"""
        <div class="stats-bar">
          <div class="stat-box"><div class="n">{len(res)}</div><div class="lb">إجمالي المنتجات</div></div>
          <div class="stat-box"><div class="n" style="color:#e53935">{len(exact_dup)}</div><div class="lb">مكرر (محذوف)</div></div>
          <div class="stat-box"><div class="n" style="color:#f9a825">{len(suspect)}</div><div class="lb">مشبوه (يحتاج مراجعة)</div></div>
          <div class="stat-box"><div class="n" style="color:#43a047">{len(new_clean)}</div><div class="lb">جديد (معتمد)</div></div>
        </div>
        """, unsafe_allow_html=True)

        # ── Suspect Products Review ────────────────────────────
        if not suspect.empty:
            st.markdown("""<div class="sec-title"><div class="bar"></div>
            <h3>المنتجات المشبوهة — راجع واعتمد أو ألغِ</h3></div>""",
                        unsafe_allow_html=True)
            st.markdown("""<div class="al-warn">
            هذه المنتجات تشبه منتجات موجودة في المتجر بنسبة عالية.
            راجع كل منتج وقرر: <b>اعتماد</b> (منتج مختلف رغم التشابه) أو <b>إلغاء</b> (مكرر).
            </div>""", unsafe_allow_html=True)

            for _, srow in suspect.iterrows():
                idx  = srow["_idx"]
                img  = srow["_img"]
                pct  = srow["نسبة التشابه"]
                approved = st.session_state.cmp_approved.get(idx, True)

                card_cls = "cmp-card suspect"
                st.markdown(f'<div class="{card_cls}">', unsafe_allow_html=True)

                cc1, cc2, cc3 = st.columns([1, 4, 2])
                with cc1:
                    if img and img.startswith("http"):
                        st.image(img, width=80)
                    else:
                        st.markdown("🖼", unsafe_allow_html=False)
                with cc2:
                    st.markdown(f"""
                    <div style="direction:rtl">
                      <div style="font-weight:800;font-size:0.95rem">{srow['الاسم الجديد']}</div>
                      <div style="color:#888;font-size:0.8rem">أقرب تطابق: {srow['أقرب تطابق في المتجر']}</div>
                      <div class="cmp-pct">{pct}% تشابه</div>
                    </div>
                    """, unsafe_allow_html=True)
                with cc3:
                    col_ap, col_cn = st.columns(2)
                    with col_ap:
                        if st.button("✅ اعتماد", key=f"ap_{idx}",
                                     type="primary" if approved else "secondary"):
                            st.session_state.cmp_approved[idx] = True
                            st.rerun()
                    with col_cn:
                        if st.button("❌ إلغاء", key=f"cn_{idx}",
                                     type="secondary" if approved else "primary"):
                            st.session_state.cmp_approved[idx] = False
                            st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

        # ── Exact Duplicates Table ─────────────────────────────
        if not exact_dup.empty:
            with st.expander(f"🔴 المنتجات المكررة ({len(exact_dup)}) — ستُحذف تلقائياً"):
                st.dataframe(exact_dup[["الاسم الجديد","SKU الجديد","أقرب تطابق في المتجر","نسبة التشابه"]],
                             use_container_width=True)

        # ── New Products Table ─────────────────────────────────
        if not new_clean.empty:
            with st.expander(f"🟢 المنتجات الجديدة المعتمدة ({len(new_clean)})"):
                st.dataframe(new_clean[["الاسم الجديد","SKU الجديد","نسبة التشابه"]],
                             use_container_width=True)

        # ── Export Final Approved List ─────────────────────────
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>تصدير القائمة النهائية المعتمدة</h3></div>""", unsafe_allow_html=True)

        if st.button("⚡ بناء الملف النهائي المعتمد", type="primary", key="build_final"):
            new_df_src = st.session_state.cmp_new_df
            approved_idxs = {idx for idx, v in st.session_state.cmp_approved.items() if v}
            # Include all "new" (auto-approved) + manually approved suspects
            final_rows = []
            for _, rrow in res.iterrows():
                idx = rrow["_idx"]
                if rrow["الحالة"] == "جديد" or \
                   (rrow["الحالة"] == "مشبوه" and st.session_state.cmp_approved.get(idx, False)):
                    if new_df_src is not None and idx in new_df_src.index:
                        final_rows.append(new_df_src.loc[idx])

            if final_rows:
                final_df = pd.DataFrame(final_rows)
                # Ensure Salla columns
                for col in SALLA_COLS:
                    if col not in final_df.columns:
                        final_df[col] = ""
                final_df = final_df[[c for c in SALLA_COLS if c in final_df.columns]]

                st.success(f"✅ {len(final_df)} منتج معتمد جاهز للرفع على سلة")
                ex1, ex2 = st.columns(2)
                with ex1:
                    st.download_button("📥 الملف النهائي — Excel",
                        export_product_xlsx(final_df),
                        f"mahwous_final_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="dl_final_x")
                with ex2:
                    st.download_button("📥 الملف النهائي — CSV",
                        export_product_csv(final_df),
                        f"mahwous_final_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        "text/csv", use_container_width=True, key="dl_final_c")

                # Also offer to move back to processor for further editing
                if st.button("🛠️ نقل للمُعالج لمزيد من التحرير", key="cmp_to_proc"):
                    st.session_state.up_df      = final_df
                    st.session_state.up_mapped  = True
                    st.session_state.up_filename = "ملف مُعتمد من المقارنة"
                    st.session_state.page       = "processor"
                    st.rerun()
            else:
                st.warning("لا توجد منتجات معتمدة — راجع قرارات الاعتماد/الإلغاء أعلاه")

        if st.button("🔄 إعادة ضبط المقارنة", key="reset_cmp"):
            st.session_state.cmp_results  = None
            st.session_state.cmp_approved = {}
            st.rerun()

    elif st.session_state.cmp_new_df is None and st.session_state.cmp_store_df is None:
        st.markdown("""
        <div class="upload-zone">
          <div class="uz-icon">🔀</div>
          <div class="uz-title">ارفع ملف المنتجات الجديدة وملف المتجر الأساسي</div>
          <div class="uz-sub">أو انقل الملف مباشرةً من المُعالج الشامل باستخدام زر "نقل للمقارنة"</div>
        </div>
        """, unsafe_allow_html=True)


# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 4 — STORE AUDIT (مدقق ملف المتجر)                        ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "store_audit":

    st.markdown("""<div class="al-info">
    ارفع ملف المتجر الأساسي (بتنسيق سلة). سيقوم النظام بفحصه واكتشاف المنتجات
    التي تحتاج معالجة (بدون صورة، بدون تصنيف، بدون ماركة، بدون وصف، بدون سعر).
    ثم يستخرج هذه المنتجات في ملف بتنسيق "ملف تحديث أو تعديل منتجات سلة" جاهز للرفع.
    </div>""", unsafe_allow_html=True)

    # ── رفع الملف ────────────────────────────────────────────────
    st.markdown("""<div class="sec-title"><div class="bar"></div><h3>رفع ملف المتجر</h3></div>""",
                unsafe_allow_html=True)

    up_audit = st.file_uploader("ارفع ملف المتجر الأساسي (CSV / Excel)",
                                 type=["csv","xlsx","xls"], key="audit_up")
    if up_audit:
        df_audit_raw = read_file(up_audit, salla_2row=True)
        if df_audit_raw.empty:
            df_audit_raw = read_file(up_audit, salla_2row=False)
        if not df_audit_raw.empty:
            st.session_state.audit_df = df_audit_raw
            st.success(f"✅ {len(df_audit_raw):,} منتج في الملف")
            st.rerun()

    if st.session_state.audit_df is not None:
        audit_df = st.session_state.audit_df

        # ── تعيين الأعمدة ────────────────────────────────────────
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>تعيين الأعمدة</h3></div>""", unsafe_allow_html=True)

        NONE_A = "— لا يوجد —"
        a_opts = [NONE_A] + list(audit_df.columns)
        def agi(kws): return a_opts.index(auto_guess_col(audit_df.columns, kws))                       if auto_guess_col(audit_df.columns, kws) in a_opts else 0

        a1, a2, a3, a4, a5, a6 = st.columns(6)
        with a1: a_no   = st.selectbox("No.", a_opts, index=agi(["no.","no","رقم","id"]), key="a_no")
        with a2: a_nm   = st.selectbox("اسم المنتج", a_opts, index=agi(["اسم","name","منتج"]), key="a_nm")
        with a3: a_img  = st.selectbox("الصورة", a_opts, index=agi(["صورة","image","img"]), key="a_img")
        with a4: a_cat  = st.selectbox("التصنيف", a_opts, index=agi(["تصنيف","category","قسم"]), key="a_cat")
        with a5: a_br   = st.selectbox("الماركة", a_opts, index=agi(["ماركة","brand","علامة"]), key="a_br")
        with a6: a_desc = st.selectbox("الوصف", a_opts, index=agi(["وصف","description","desc"]), key="a_desc")

        a7, a8 = st.columns(2)
        with a7: a_pr   = st.selectbox("السعر", a_opts, index=agi(["سعر","price"]), key="a_pr")
        with a8: a_sku  = st.selectbox("SKU", a_opts, index=agi(["sku","رمز","barcode"]), key="a_sku")

        # ── تشغيل الفحص ──────────────────────────────────────────
        if st.button("🔍 فحص الملف الآن", type="primary", key="run_audit"):
            issues = []
            for i, row in audit_df.iterrows():
                row_issues = []
                name = str(row.get(a_nm, "") or "").strip() if a_nm != NONE_A else ""
                if not name or name == "nan":
                    continue

                if a_img != NONE_A:
                    img_val = str(row.get(a_img, "") or "").strip()
                    if not img_val or img_val == "nan":
                        row_issues.append("بدون صورة")

                if a_cat != NONE_A:
                    cat_val = str(row.get(a_cat, "") or "").strip()
                    if not cat_val or cat_val == "nan":
                        row_issues.append("بدون تصنيف")

                if a_br != NONE_A:
                    br_val = str(row.get(a_br, "") or "").strip()
                    if not br_val or br_val == "nan":
                        row_issues.append("بدون ماركة")

                if a_desc != NONE_A:
                    desc_val = str(row.get(a_desc, "") or "").strip()
                    if not desc_val or desc_val == "nan" or len(desc_val) < 20:
                        row_issues.append("بدون وصف")

                if a_pr != NONE_A:
                    pr_val = str(row.get(a_pr, "") or "").strip()
                    if not pr_val or pr_val in ["0", "nan", ""]:
                        row_issues.append("بدون سعر")

                if row_issues:
                    issues.append({
                        "No.":               str(row.get(a_no, i) or i) if a_no != NONE_A else str(i),
                        "أسم المنتج":        name,
                        "الماركة":           str(row.get(a_br, "") or "") if a_br != NONE_A else "",
                        "تصنيف المنتج":      str(row.get(a_cat, "") or "") if a_cat != NONE_A else "",
                        "صورة المنتج":       str(row.get(a_img, "") or "") if a_img != NONE_A else "",
                        "وصف صورة المنتج":   name,
                        "نوع المنتج":        "منتج جاهز",
                        "سعر المنتج":        str(row.get(a_pr, "") or "") if a_pr != NONE_A else "",
                        "الوصف":             str(row.get(a_desc, "") or "") if a_desc != NONE_A else "",
                        "هل يتطلب شحن؟":    "نعم",
                        "رمز المنتج sku":    str(row.get(a_sku, "") or "") if a_sku != NONE_A else "",
                        "الوزن":             "0.2",
                        "وحدة الوزن":        "kg",
                        "خاضع للضريبة ؟":   "نعم",
                        "اقصي كمية لكل عميل": "0",
                        "تثبيت المنتج":      "لا",
                        "اضافة صورة عند الطلب": "لا",
                        "_issues":           " | ".join(row_issues),
                        "_idx":              i,
                    })

            st.session_state.audit_results = pd.DataFrame(issues) if issues else pd.DataFrame()
            st.rerun()

        # ── عرض نتائج الفحص ──────────────────────────────────────
        if st.session_state.audit_results is not None:
            audit_res = st.session_state.audit_results

            if audit_res.empty:
                st.success("✅ الملف مكتمل — لا توجد منتجات تحتاج معالجة!")
            else:
                # إحصائيات
                no_img  = int(audit_res["_issues"].str.contains("بدون صورة").sum())
                no_cat  = int(audit_res["_issues"].str.contains("بدون تصنيف").sum())
                no_br   = int(audit_res["_issues"].str.contains("بدون ماركة").sum())
                no_desc = int(audit_res["_issues"].str.contains("بدون وصف").sum())
                no_pr   = int(audit_res["_issues"].str.contains("بدون سعر").sum())

                st.markdown(f"""
                <div class="stats-bar">
                  <div class="stat-box"><div class="n" style="color:#e53935">{len(audit_res):,}</div><div class="lb">تحتاج معالجة</div></div>
                  <div class="stat-box"><div class="n" style="color:#f9a825">{no_img:,}</div><div class="lb">بدون صورة</div></div>
                  <div class="stat-box"><div class="n" style="color:#f9a825">{no_cat:,}</div><div class="lb">بدون تصنيف</div></div>
                  <div class="stat-box"><div class="n" style="color:#f9a825">{no_br:,}</div><div class="lb">بدون ماركة</div></div>
                  <div class="stat-box"><div class="n" style="color:#f9a825">{no_desc:,}</div><div class="lb">بدون وصف</div></div>
                  <div class="stat-box"><div class="n" style="color:#f9a825">{no_pr:,}</div><div class="lb">بدون سعر</div></div>
                </div>
                """, unsafe_allow_html=True)

                # فلتر حسب نوع المشكلة
                filter_opts = ["الكل", "بدون صورة", "بدون تصنيف", "بدون ماركة", "بدون وصف", "بدون سعر"]
                audit_filter = st.selectbox("فلتر حسب المشكلة:", filter_opts, key="audit_filter")

                if audit_filter == "الكل":
                    filtered_audit = audit_res
                else:
                    filtered_audit = audit_res[audit_res["_issues"].str.contains(audit_filter)]

                # عرض الجدول
                display_cols = ["No.", "أسم المنتج", "الماركة", "تصنيف المنتج", "_issues"]
                st.dataframe(
                    filtered_audit[[c for c in display_cols if c in filtered_audit.columns]],
                    use_container_width=True
                )

                # تصدير بتنسيق ملف تحديث سلة
                st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
                <h3>تصدير للمعالجة</h3></div>""", unsafe_allow_html=True)

                st.info(f"سيتم تصدير {len(filtered_audit):,} منتج بتنسيق 'ملف تحديث أو تعديل منتجات سلة' جاهز للرفع بعد المعالجة.")

                # بناء ملف التحديث بتنسيق سلة
                def build_update_file(df_issues: pd.DataFrame) -> pd.DataFrame:
                    update_cols = SALLA_COLS
                    rows = []
                    for _, r in df_issues.iterrows():
                        row = {col: "" for col in update_cols}
                        for col in update_cols:
                            if col in r.index:
                                row[col] = str(r[col] or "")
                        row["النوع "] = "منتج"
                        rows.append(row)
                    return pd.DataFrame(rows, columns=update_cols)

                update_df = build_update_file(filtered_audit)

                aud_e1, aud_e2, aud_e3 = st.columns(3)
                with aud_e1:
                    st.download_button(
                        f"📥 ملف التحديث — Excel ({len(update_df):,} منتج)",
                        export_product_xlsx(update_df),
                        f"تحديث_منتجات_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key="dl_audit_x"
                    )
                with aud_e2:
                    st.download_button(
                        f"📥 ملف التحديث — CSV ({len(update_df):,} منتج)",
                        export_product_csv(update_df),
                        f"تحديث_منتجات_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        "text/csv", use_container_width=True, key="dl_audit_c"
                    )
                with aud_e3:
                    if st.button("🛠️ نقل للمُعالج لإكمال البيانات", key="audit_to_proc",
                                 use_container_width=True):
                        st.session_state.up_df      = update_df
                        st.session_state.up_mapped  = True
                        st.session_state.up_filename = f"تحديث_منتجات_{len(update_df)}"
                        st.session_state.page       = "processor"
                        st.rerun()

                if st.button("🔄 إعادة الفحص", key="reset_audit"):
                    st.session_state.audit_results = None
                    st.rerun()

    else:
        st.markdown("""
        <div class="upload-zone">
          <div class="uz-icon">🏪</div>
          <div class="uz-title">ارفع ملف المتجر الأساسي للبدء</div>
          <div class="uz-sub">سيكتشف النظام المنتجات الناقصة ويجهّز ملف التحديث تلقائياً</div>
        </div>
        """, unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 5 — BRANDS CHECKER                                        ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "brands":

    st.markdown("""<div class="al-info">
    ارفع قائمة ماركات للتحقق منها مقابل قاعدة بيانات مهووس.
    المنتجات التي لا تجد لها ماركة ستُولَّد لها ماركة جديدة بتنسيق سلة.
    </div>""", unsafe_allow_html=True)

    up_b = st.file_uploader("ارفع ملف الماركات أو المنتجات",
                             type=["csv","xlsx","xls"], key="brands_up")
    if up_b:
        bdf_raw = read_file(up_b)
        if not bdf_raw.empty:
            st.success(f"✅ {len(bdf_raw)} صف")
            with st.expander("👀 معاينة"): st.dataframe(bdf_raw.head(8), use_container_width=True)

            NONE_B = "— لا يوجد —"
            bopts  = [NONE_B] + list(bdf_raw.columns)
            def bgi(kws): return bopts.index(auto_guess_col(bdf_raw.columns, kws)) \
                          if auto_guess_col(bdf_raw.columns, kws) in bopts else 0

            b1, b2 = st.columns(2)
            with b1: bcol_name = st.selectbox("عمود اسم الماركة:", bopts,
                                               index=bgi(["ماركة","brand","علامة","اسم"]), key="bcol_nm")
            with b2: bcol_prod = st.selectbox("عمود اسم المنتج (اختياري):", bopts,
                                               index=bgi(["منتج","product","اسم"]), key="bcol_pr")

            gen_missing = st.checkbox("🤖 توليد ماركات جديدة بالذكاء الاصطناعي للماركات الغير موجودة",
                                       value=True, key="gen_miss_b")

            if st.button("🔍 تدقيق الآن", type="primary", key="check_brands"):
                if bcol_name == NONE_B:
                    st.error("حدد عمود اسم الماركة")
                else:
                    results_b = []
                    new_brands_b = []
                    for _, row in bdf_raw.iterrows():
                        bname = str(row.get(bcol_name, "") or "").strip()
                        pname = str(row.get(bcol_prod, "") or "").strip() \
                                if bcol_prod != NONE_B else ""
                        if not bname and not pname:
                            continue
                        search_name = bname or pname
                        found = match_brand(search_name)
                        status = "موجودة ✅" if found.get("name") else "غير موجودة ❌"
                        results_b.append({
                            "الماركة المدخلة":  bname,
                            "المنتج":           pname,
                            "الماركة في قاعدة البيانات": found.get("name","—"),
                            "الرابط":           found.get("page_url",""),
                            "الحالة":           status,
                        })
                        if not found.get("name") and bname:
                            existing_new = [b["اسم العلامة التجارية"] for b in st.session_state.new_brands]
                            if bname not in existing_new and bname not in [x["اسم العلامة التجارية"] for x in new_brands_b]:
                                new_brands_b.append({
                                    "اسم العلامة التجارية": bname,
                                    "(SEO Page URL) رابط صفحة العلامة التجارية": to_slug(bname),
                                    "وصف العلامة التجارية": "",
                                    "صورة العلامة التجارية": "",
                                })

                    res_b_df = pd.DataFrame(results_b)
                    st.dataframe(res_b_df, use_container_width=True)

                    found_c   = int((res_b_df["الحالة"] == "موجودة ✅").sum())
                    missing_c = int((res_b_df["الحالة"] == "غير موجودة ❌").sum())
                    st.markdown(f"""
                    <div class="stats-bar">
                      <div class="stat-box"><div class="n" style="color:#43a047">{found_c}</div><div class="lb">موجودة</div></div>
                      <div class="stat-box"><div class="n" style="color:#e53935">{missing_c}</div><div class="lb">غير موجودة</div></div>
                    </div>
                    """, unsafe_allow_html=True)

                    if new_brands_b:
                        if gen_missing and st.session_state.api_key:
                            with st.spinner(f"توليد أوصاف {len(new_brands_b)} ماركة جديدة..."):
                                for nb in new_brands_b:
                                    gen = generate_new_brand(nb["اسم العلامة التجارية"])
                                    nb["وصف العلامة التجارية"] = gen["وصف العلامة التجارية"]
                        st.session_state.new_brands.extend(new_brands_b)
                        st.info(f"🆕 {len(new_brands_b)} ماركة جديدة أُضيفت — صدّرها من الإعدادات")

    # Show pending new brands
    if st.session_state.new_brands:
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>الماركات الجديدة المُولَّدة</h3></div>""", unsafe_allow_html=True)
        nb_df = pd.DataFrame(st.session_state.new_brands)
        edited_nb = st.data_editor(nb_df, use_container_width=True,
                                    num_rows="dynamic", key="nb_editor")
        st.session_state.new_brands = edited_nb.to_dict("records")

        nb1, nb2, nb3 = st.columns(3)
        with nb1:
            st.download_button("📥 ماركات جديدة — Excel",
                export_brands_xlsx(st.session_state.new_brands),
                "new_brands.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_nb_x")
        with nb2:
            nb_csv_buf = io.StringIO()
            nb_csv_buf.write(",".join(SALLA_BRANDS_COLS) + "\n")
            for nb in st.session_state.new_brands:
                nb_csv_buf.write(",".join([f'"{str(nb.get(c,"") or "")}"'
                                           for c in SALLA_BRANDS_COLS]) + "\n")
            st.download_button("📥 ماركات جديدة — CSV",
                nb_csv_buf.getvalue().encode("utf-8-sig"),
                "new_brands.csv", "text/csv",
                use_container_width=True, key="dl_nb_c")
        with nb3:
            if st.button("🗑️ مسح الماركات الجديدة", key="clear_nb"):
                st.session_state.new_brands = []
                st.rerun()

        # Generate AI descriptions for brands without descriptions
        no_desc_brands = [b for b in st.session_state.new_brands
                          if not str(b.get("وصف العلامة التجارية","")).strip()]
        if no_desc_brands and st.session_state.api_key:
            if st.button(f"🤖 توليد أوصاف {len(no_desc_brands)} ماركة بدون وصف", key="gen_nb_desc"):
                with st.spinner("توليد الأوصاف..."):
                    for nb in st.session_state.new_brands:
                        if not str(nb.get("وصف العلامة التجارية","")).strip():
                            gen = generate_new_brand(nb["اسم العلامة التجارية"])
                            nb["وصف العلامة التجارية"] = gen["وصف العلامة التجارية"]
                st.success("✅ تم توليد الأوصاف")
                st.rerun()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 6 — SETTINGS                                              ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "settings":

    st.markdown("""<div class="sec-title"><div class="bar"></div>
    <h3>مفاتيح API</h3></div>""", unsafe_allow_html=True)

    s1, s2 = st.columns(2)
    with s1:
        new_key = st.text_input("🔑 Anthropic API Key (Claude):",
                                value=st.session_state.api_key,
                                type="password", key="set_api")
        if st.button("💾 حفظ مفتاح Claude", key="save_api"):
            st.session_state.api_key = new_key
            st.success("✅ تم حفظ المفتاح")
    with s2:
        new_gk = st.text_input("🔑 Google API Key:",
                               value=st.session_state.google_api,
                               type="password", key="set_gk")
        new_cx = st.text_input("🔍 Google CSE ID:",
                               value=st.session_state.google_cse,
                               type="password", key="set_cx")
        if st.button("💾 حفظ مفاتيح Google", key="save_gk"):
            st.session_state.google_api = new_gk
            st.session_state.google_cse = new_cx
            st.success("✅ تم حفظ المفاتيح")

    st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
    <h3>قواعد البيانات المرجعية</h3></div>""", unsafe_allow_html=True)

    db1, db2 = st.columns(2)
    with db1:
        st.markdown("**ملف الماركات**")
        bdf = st.session_state.brands_df
        if bdf is not None:
            st.markdown(f'<div class="al-ok">{len(bdf)} ماركة محملة</div>',
                        unsafe_allow_html=True)
            with st.expander("👀 معاينة"): st.dataframe(bdf.head(5), use_container_width=True)
        up_brands = st.file_uploader("تحديث ملف الماركات:", type=["csv","xlsx"],
                                      key="up_brands_db")
        if up_brands:
            df_b = read_file(up_brands)
            if not df_b.empty:
                st.session_state.brands_df = df_b
                os.makedirs(DATA_DIR, exist_ok=True)
                df_b.to_csv(os.path.join(DATA_DIR, "brands.csv"),
                            index=False, encoding="utf-8-sig")
                st.success(f"✅ تم تحديث {len(df_b)} ماركة")
                st.rerun()

    with db2:
        st.markdown("**ملف التصنيفات**")
        cdf = st.session_state.categories_df
        if cdf is not None:
            st.markdown(f'<div class="al-ok">{len(cdf)} تصنيف محمّل</div>',
                        unsafe_allow_html=True)
            with st.expander("👀 معاينة"): st.dataframe(cdf.head(5), use_container_width=True)
        up_cats = st.file_uploader("تحديث ملف التصنيفات:", type=["csv","xlsx"],
                                    key="up_cats_db")
        if up_cats:
            df_c = read_file(up_cats)
            if not df_c.empty:
                st.session_state.categories_df = df_c
                os.makedirs(DATA_DIR, exist_ok=True)
                df_c.to_csv(os.path.join(DATA_DIR, "categories.csv"),
                            index=False, encoding="utf-8-sig")
                st.success(f"✅ تم تحديث {len(df_c)} تصنيف")
                st.rerun()

    # New brands export section
    if st.session_state.new_brands:
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>الماركات الجديدة المُولَّدة — جاهزة للتصدير</h3></div>""", unsafe_allow_html=True)
        st.markdown(f'<div class="al-warn">{len(st.session_state.new_brands)} ماركة جديدة اكتُشفت خلال المعالجة وتحتاج إلى إضافتها لمتجرك على سلة.</div>',
                    unsafe_allow_html=True)
        nb_df_s = pd.DataFrame(st.session_state.new_brands)
        st.dataframe(nb_df_s, use_container_width=True)
        sn1, sn2 = st.columns(2)
        with sn1:
            st.download_button("📥 تصدير الماركات الجديدة — Excel",
                export_brands_xlsx(st.session_state.new_brands),
                "new_brands_salla.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="exp_nb_set_x")
        with sn2:
            nb_csv_s = io.StringIO()
            nb_csv_s.write(",".join(SALLA_BRANDS_COLS) + "\n")
            for nb in st.session_state.new_brands:
                nb_csv_s.write(",".join([f'"{str(nb.get(c,"") or "")}"'
                                          for c in SALLA_BRANDS_COLS]) + "\n")
            st.download_button("📥 تصدير الماركات الجديدة — CSV",
                nb_csv_s.getvalue().encode("utf-8-sig"),
                "new_brands_salla.csv", "text/csv",
                use_container_width=True, key="exp_nb_set_c")

    st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
    <h3>معلومات النظام</h3></div>""", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="direction:rtl;font-size:0.85rem;line-height:2">
      <b>الإصدار:</b> مهووس مركز التحكم الشامل v4.8<br>
      <b>الموقع:</b> <a href="https://mahwous-automation-production.up.railway.app/" target="_blank">mahwous-automation-production.up.railway.app</a><br>
      <b>أعمدة سلة المنتجات:</b> {len(SALLA_COLS)} عمود<br>
      <b>أعمدة سلة SEO:</b> {len(SALLA_SEO_COLS)} عمود<br>
      <b>أعمدة تحديث الأسعار:</b> {len(SALLA_PRICE_COLS)} عمود<br>
      <b>أعمدة ملف الماركات:</b> {len(SALLA_BRANDS_COLS)} عمود<br>
    </div>
    """, unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  FOOTER                                                         ║
# ╚══════════════════════════════════════════════════════════════════╝
st.markdown("""
<div class="mhw-footer">
  مهووس — مركز التحكم الشامل v4.8 &nbsp;|&nbsp;
  جميع الملفات المُصدَّرة متوافقة 100% مع منصة سلة &nbsp;|&nbsp;
  <a href="https://mahwous-automation-production.up.railway.app/" target="_blank">mahwous.com</a>
</div>
""", unsafe_allow_html=True)
